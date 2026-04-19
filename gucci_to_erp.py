#!/usr/bin/env python3
"""
Gucci 图片链接表 → ERP 标准导入模板转换器

关键约束 (P0):
  - I列必须 N (下架)
  - 单商品图片总数 ≤ 14 (含首图)
  - E列单URL, F列换行分隔多URL
  - P列重量为 kg，3位小数
  - AD/AE 为 2位小数
  - J列物流模板必须是系统已配置模板

输出遵循 templates/商品导入模板 (修改版1.0).xlsx 的结构与样式，
每个商品生成 4 行 SKU (Size:S/M/L/XL)。
"""

import argparse
from copy import copy
import re
import sys
from pathlib import Path
from typing import DefaultDict, Iterable, List, Optional, Tuple
from collections import Counter, defaultdict

import openpyxl

COL_ID       = 1   # A
COL_TITLE    = 2   # B
COL_SUBTITLE = 3   # C
COL_DESC     = 4   # D
COL_MAIN_IMG = 5   # E
COL_OTHER_IMG= 6   # F
COL_KEYINFO  = 7   # G
COL_ATTR     = 8   # H
COL_PUBLISH  = 9   # I
COL_LOGISTICS= 10  # J
COL_CATEGORY = 11  # K
COL_TAGS     = 12  # L
COL_UNIT     = 13  # M
COL_NOTE     = 14  # N
COL_NO_INV   = 15  # O
COL_WEIGHT   = 16  # P
COL_LENGTH   = 17  # Q
COL_WIDTH    = 18  # R
COL_HEIGHT   = 19  # S
COL_SEO_TITLE= 20  # T
COL_SEO_DESC = 21  # U
COL_SEO_KEY  = 22  # V
COL_SEO_URL  = 23  # W
COL_SPEC1    = 24  # X
COL_SPEC2    = 25  # Y
COL_SPEC3    = 26  # Z
COL_SPEC4    = 27  # AA
COL_SKU_VAL  = 28  # AB
COL_SKU_IMG  = 29  # AC
COL_PRICE    = 30  # AD
COL_ORIG_PRICE= 31 # AE
COL_INVENTORY= 32  # AF
COL_SKU_CODE = 33  # AG

SIZE_LIST = ["S", "M", "L", "XL"]


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub("[^a-z0-9\\s-]", "", text)           # keep alnum, whitespace, hyphen
    text = re.sub("[\\s]+", "-", text)                 # spaces → hyphens
    text = re.sub("-+", "-", text).strip("-")
    return f"gucci-{text}" if text else "gucci-item"


def _escape_html(s: object) -> str:
    s = "" if s is None else str(s)
    s = s.replace("&", "&amp;")
    s = s.replace("<", "&lt;")
    s = s.replace(">", "&gt;")
    s = s.replace('"', "&quot;")
    s = s.replace("'", "&apos;")
    return s


def _strip_brand(title: str, brand: str) -> str:
    t = str(title or "").strip()
    b = str(brand or "").strip()
    if not t or not b:
        return t
    return re.sub(rf"^{re.escape(b)}\s*[/]?\s*", "", t, flags=re.IGNORECASE).strip()


def _build_desc_html(title: str, brand: str) -> str:
    name_field = _strip_brand(title, brand)
    brand_hyphen = brand.replace(" ", "-")
    html = (
        f'<p><span style="font-family: Tahoma;"><span>Name: '
        f'<span style="font-family: verdana, geneva, sans-serif;">'
        f'<a href="https://www.stockxshoesvip.net/{brand_hyphen}/" '
        f'rel="noopener" target="_blank">{_escape_html(brand)}</a> {_escape_html(name_field)}'
        f'</span></span></span></p>'
        f'<p>Category: <span style="font-size: 14px;">'
        f'<a href="https://www.stockxshoesvip.net/{brand_hyphen}-Clothes/" '
        f'target="_self" class="third-link animation-underline">{_escape_html(brand)} Clothes</a>'
        f'</span></p>'
        f'<p><b>Our Core Guarantees</b></p>'
        f'<ul>'
        f'<li><b>Exclusive <a href="https://www.stockxshoesvip.net/Stockxshoes-QC-Pics/" '
        f'rel="noopener" target="_blank">QC Service</a>:</b> We provide free Quality Control (QC) '
        f'pictures before shipment. You approve the exact item you will receive.</li>'
        f'<li><b>Premium Packaging:</b> All apparel comes with full brand packaging and original tags.</li>'
        f'<li><b>Worry-Free Logistics:</b> Secure delivery and customs clearance.</li>'
        f'<li><b>100% Safe Shopping:</b> 30-day money-back guarantee.</li>'
        f'</ul>'
        f'<p><b>Shipping &amp; Payment</b></p>'
        f'<ul>'
        f'<li><b>Delivery Time:</b> 7-18 Days. Tracking number provided.</li>'
        f'<li><b>Shipping Methods:</b> FedEx / USPS / DHL / UPS / EMS / Royal Mail.</li>'
        f'<li><b>Payment Methods:</b> Credit/Debit Card, PayPal, Bank Transfer, Cash App, Zelle.</li>'
        f'</ul>'
        f'<p><b>About StockxShoesVIP</b></p>'
        f'<p>With 10 years of offline retail and 5 years of online excellence.</p>'
        f'<p><i>(Note: We are not affiliated with the StockX platform. '
        f'Please bookmark: stockxshoesvip.net)</i></p>'
        f'<p><b>Contact Us</b></p>'
        f'<ul>'
        f'<li><b>WhatsApp/WeChat:</b> +86 189 5920 5893</li>'
        f'<li><b>Instagram:</b> @stockxshoesvip_com</li>'
        f'</ul>'
        f'<p><i>Buy with confidence, wear with confidence.</i></p>'
    )
    return html


def _build_seo_title(title: str) -> str:
    return f"Stockx Replica Streetwear | Top Quality 1:1 {title} - stockxshoesvip.net"


def _build_seo_desc(title: str) -> str:
    return (
        "Buy Best 1:1 Replica Clothing on Stockxshoesvip.net. "
        f"Perfect {title}. 100% safe shipping, free QC confirmation, and easy returns."
    )


def _resource_root() -> Path:
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return Path(meipass)
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _clear_row_values(ws, row_idx: int, max_col: int = 33) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row_idx, c).value = None


def _make_unique_title(title: str, used_titles: set, seq: Optional[int] = None) -> str:
    base = str(title or "").strip()
    if not base:
        return base

    max_len = 255
    if seq is None:
        candidate = base[:max_len]
    else:
        suffix = f" {seq:02d}"
        keep_len = max_len - len(suffix)
        candidate = f"{base[:keep_len].rstrip()}{suffix}"

    ckey = candidate.lower()
    if ckey not in used_titles:
        used_titles.add(ckey)
        return candidate

    idx = 2
    while True:
        if seq is None:
            suffix = f" ({idx})"
        else:
            suffix = f" {seq:02d}-{idx}"
        keep_len = max_len - len(suffix)
        candidate2 = f"{base[:keep_len].rstrip()}{suffix}"
        ckey2 = candidate2.lower()
        if ckey2 not in used_titles:
            used_titles.add(ckey2)
            return candidate2
        idx += 1


def _is_url(val: object) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return s.startswith("http://") or s.startswith("https://")


def _uniq_preserve_order(items: Iterable[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for it in items:
        if it in seen:
            continue
        seen.add(it)
        out.append(it)
    return out


def _extract_images(src_row: Tuple[object, ...]) -> Tuple[str, List[str]]:
    main_img = str(src_row[3]).strip() if len(src_row) > 3 and _is_url(src_row[3]) else ""
    other_imgs: List[str] = []
    for v in src_row[4:]:
        if _is_url(v):
            other_imgs.append(str(v).strip())

    other_imgs = _uniq_preserve_order(other_imgs)
    if main_img:
        other_imgs = [u for u in other_imgs if u != main_img]
    elif other_imgs:
        main_img = other_imgs.pop(0)

    max_total = 14
    other_limit = max_total - 1
    if len(other_imgs) > other_limit:
        other_imgs = other_imgs[:other_limit]

    return main_img, other_imgs


def _snapshot_row(ws, row_idx: int, max_col: int = 33) -> dict:
    row_dim = ws.row_dimensions[row_idx]
    cells = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        cells.append(
            {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
                "style": cell.style,
            }
        )
    return {"height": row_dim.height, "cells": cells}


def _apply_row_snapshot(ws, row_idx: int, snap: dict) -> None:
    ws.row_dimensions[row_idx].height = snap["height"]
    cells = snap["cells"]
    for c in range(1, len(cells) + 1):
        cell = ws.cell(row=row_idx, column=c)
        s = cells[c - 1]
        cell.font = copy(s["font"])
        cell.fill = copy(s["fill"])
        cell.border = copy(s["border"])
        cell.alignment = copy(s["alignment"])
        cell.number_format = s["number_format"]
        cell.protection = copy(s["protection"])
        cell.style = s["style"]


def build_erp_xlsx(
    src_path: Path,
    out_path: Path,
    tpl_path: Path,
    brand: str,
    logistics_template: str,
    attribute: str,
    unit: str,
    weight_kg: float,
    price: float,
    original_price: float,
    stock: int,
) -> None:
    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = wb_src.active
    src_rows = list(ws_src.iter_rows(values_only=True))
    wb_src.close()

    if len(src_rows) <= 1:
        raise ValueError("源文件没有可转换的数据行")

    wb_out = openpyxl.load_workbook(tpl_path)
    ws_out = wb_out["商品信息 (2)"] if "商品信息 (2)" in wb_out.sheetnames else wb_out.active

    base_snaps = [
        _snapshot_row(ws_out, 4, 33),
        _snapshot_row(ws_out, 5, 33),
        _snapshot_row(ws_out, 6, 33),
        _snapshot_row(ws_out, 7, 33),
    ]

    if ws_out.max_row >= 4:
        ws_out.delete_rows(4, ws_out.max_row - 3)

    products = src_rows[1:]
    total_rows = len(products) * 4
    ws_out.insert_rows(4, amount=total_rows)

    errors: List[str] = []
    used_titles: set = set()
    title_counts = Counter()
    for src_row in products:
        t = str(src_row[1]).strip() if len(src_row) > 1 and src_row[1] is not None else ""
        if t:
            title_counts[t.lower()] += 1

    title_seq: DefaultDict[str, int] = defaultdict(int)
    for i, src_row in enumerate(products, start=1):
        title = str(src_row[1]).strip() if len(src_row) > 1 and src_row[1] is not None else ""
        if not title:
            errors.append(f"第{i}行: 商品标题为空")
            continue

        main_img, other_imgs = _extract_images(src_row)
        if not main_img:
            errors.append(f"第{i}行: 商品首图为空")
            continue

        other_imgs_str = "\n".join(other_imgs) if other_imgs else ""
        group_start = 4 + (i - 1) * 4
        for j in range(4):
            _apply_row_snapshot(ws_out, group_start + j, base_snaps[j])

        r0 = group_start
        tkey = title.lower()
        seq = None
        if title_counts.get(tkey, 0) > 1:
            title_seq[tkey] += 1
            seq = title_seq[tkey]
        unique_title = _make_unique_title(title, used_titles, seq=seq)
        desc_html = _build_desc_html(unique_title, brand)
        seo_title = _build_seo_title(unique_title)
        seo_desc = _build_seo_desc(unique_title)
        for rr in range(group_start, group_start + 4):
            _clear_row_values(ws_out, rr, 33)

        ws_out.cell(r0, COL_TITLE).value = unique_title
        ws_out.cell(r0, COL_DESC).value = desc_html
        ws_out.cell(r0, COL_MAIN_IMG).value = main_img
        ws_out.cell(r0, COL_OTHER_IMG).value = other_imgs_str
        ws_out.cell(r0, COL_ATTR).value = attribute
        ws_out.cell(r0, COL_PUBLISH).value = "N"
        ws_out.cell(r0, COL_LOGISTICS).value = logistics_template
        ws_out.cell(r0, COL_CATEGORY).value = brand
        ws_out.cell(r0, COL_TAGS).value = unique_title
        ws_out.cell(r0, COL_UNIT).value = unit
        ws_out.cell(r0, COL_NO_INV).value = "Y"
        ws_out.cell(r0, COL_WEIGHT).value = float(weight_kg)
        ws_out.cell(r0, COL_SEO_TITLE).value = seo_title
        ws_out.cell(r0, COL_SEO_DESC).value = seo_desc
        ws_out.cell(r0, COL_SEO_KEY).value = unique_title
        ws_out.cell(r0, COL_SEO_URL).value = None
        ws_out.cell(r0, COL_SPEC2).value = "Size\nS\nM\nL\nXL"

        for j, size in enumerate(SIZE_LIST):
            rr = group_start + j
            ws_out.cell(rr, COL_SKU_VAL).value = f"Size:{size}"
            ws_out.cell(rr, COL_PRICE).value = float(price)
            ws_out.cell(rr, COL_ORIG_PRICE).value = float(original_price)
            ws_out.cell(rr, COL_INVENTORY).value = int(stock)

    try:
        wb_out.save(out_path)
    except PermissionError as e:
        raise PermissionError(f"无法写入输出文件(可能正在被Excel/WPS占用): {out_path}") from e
    if errors:
        raise ValueError("转换失败，存在不合规数据:\n" + "\n".join(errors))


def main(argv: Optional[List[str]] = None) -> int:
    res_root = _resource_root()
    default_src = Path.cwd() / "67款GucciT恤图片链接表_0419.xlsx"
    default_out = Path.cwd() / "GucciT恤_ERP导入模板_0419_BAPE标准对齐_标题去重.xlsx"
    default_tpl = res_root / "templates" / "商品导入模板 (修改版1.0).xlsx"

    p = argparse.ArgumentParser(description="Gucci 图片链接表 → ERP 标准导入模板 (.xlsx)")
    p.add_argument("--src", type=Path, default=default_src, help="源文件路径 (.xlsx)")
    p.add_argument("--out", type=Path, default=default_out, help="输出文件路径 (.xlsx)")
    p.add_argument("--tpl", type=Path, default=default_tpl, help="ERP模板路径 (.xlsx)")
    p.add_argument("--brand", type=str, default="Gucci", help="类别/品牌名 (写入K列)")
    p.add_argument("--logistics", type=str, default="Clothing", help="物流模板 (写入J列)")
    p.add_argument("--attribute", type=str, default="材质|棉质", help="属性 (写入H列)")
    p.add_argument("--unit", type=str, default="件/个", help="计量单位 (写入M列)")
    p.add_argument("--weight", type=float, default=0.3, help="重量kg (写入P列)")
    p.add_argument("--price", type=float, default=59.0, help="售价 (写入AD列)")
    p.add_argument("--original-price", type=float, default=99.0, help="原价 (写入AE列)")
    p.add_argument("--stock", type=int, default=999, help="库存 (写入AF列)")
    args = p.parse_args(argv)

    if not args.src.exists():
        raise FileNotFoundError(f"源文件不存在: {args.src}")
    if not args.tpl.exists():
        raise FileNotFoundError(f"模板不存在: {args.tpl}")

    build_erp_xlsx(
        src_path=args.src,
        out_path=args.out,
        tpl_path=args.tpl,
        brand=args.brand,
        logistics_template=args.logistics,
        attribute=args.attribute,
        unit=args.unit,
        weight_kg=args.weight,
        price=args.price,
        original_price=args.original_price,
        stock=args.stock,
    )

    print(str(args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
