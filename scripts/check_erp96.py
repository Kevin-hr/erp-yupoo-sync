# -*- coding: utf-8 -*-
import openpyxl, warnings, re, sys
warnings.filterwarnings('ignore')

fpath = r'C:\Users\Administrator\Documents\GitHub\erp-yupoo-sync\inputs\ERP_Export_2026-04-28_96products_final.xlsx'
wb = openpyxl.load_workbook(fpath, data_only=True)
ws = wb.active

def col_letter(n):
    return openpyxl.utils.get_column_letter(n)

print('=== 所有列头（row=1）===')
for cell in ws[1]:
    if cell.value is not None and str(cell.value).strip():
        print('  ' + cell.coordinate + ': ' + repr(str(cell.value)))

print()
print('=== 主行(Size:S)关键字段 ===')
for row_idx in range(1, ws.max_row + 1):
    ab_val = str(ws.cell(row=row_idx, column=28).value or '').strip()
    if ab_val == 'Size:S':
        b  = str(ws.cell(row=row_idx, column=2).value or '')
        d  = str(ws.cell(row=row_idx, column=4).value or '')
        e  = str(ws.cell(row=row_idx, column=5).value or '')
        f  = str(ws.cell(row=row_idx, column=6).value or '')
        h  = str(ws.cell(row=row_idx, column=8).value or '')
        i_ = str(ws.cell(row=row_idx, column=9).value or '')
        j  = str(ws.cell(row=row_idx, column=10).value or '')
        k  = str(ws.cell(row=row_idx, column=11).value or '')
        l  = str(ws.cell(row=row_idx, column=12).value or '')
        m  = str(ws.cell(row=row_idx, column=13).value or '')
        o  = str(ws.cell(row=row_idx, column=15).value or '')
        p  = str(ws.cell(row=row_idx, column=16).value or '')
        t  = str(ws.cell(row=row_idx, column=20).value or '')
        u  = str(ws.cell(row=row_idx, column=21).value or '')
        v  = str(ws.cell(row=row_idx, column=22).value or '')
        y  = str(ws.cell(row=row_idx, column=25).value or '')
        ad = ws.cell(row=row_idx, column=30).value
        ae = ws.cell(row=row_idx, column=31).value
        af = ws.cell(row=row_idx, column=32).value

        print('  B=' + repr(b[:40]))
        print('  D[HTML前60]=' + repr(d[:60]))
        print('  E=' + repr(e[:60]))
        f_count = f.count('\n')+1 if f.startswith('http') else 0
        print('  F图片数=' + str(f_count))
        print('  H=' + repr(h) + ' I=' + repr(i_) + ' J=' + repr(j) + ' K=' + repr(k))
        print('  L=' + repr(l[:40]) + ' M=' + repr(m) + ' O=' + repr(o) + ' P=' + repr(p))
        print('  T=' + repr(t[:80]))
        print('  U=' + repr(u[:80]))
        print('  V=' + repr(v[:40]))
        print('  Y=' + repr(y[:40]))
        print('  AB=' + repr(ab_val) + ' AD=' + repr(ad) + ' AE=' + repr(ae) + ' AF=' + repr(af))
        break

print()
print('=== 字段合规性汇总 ===')
issues = []
for row_idx in range(1, ws.max_row + 1):
    ab_val = str(ws.cell(row=row_idx, column=28).value or '').strip()
    if ab_val != 'Size:S': continue

    i_val = str(ws.cell(row=row_idx, column=9).value or '').strip()
    if i_val != 'N': issues.append('Row' + str(row_idx) + ' I=' + repr(i_val) + ' [标准=N]')

    j_val = str(ws.cell(row=row_idx, column=10).value or '').strip()
    if j_val != 'Clothing': issues.append('Row' + str(row_idx) + ' J=' + repr(j_val) + ' [标准=Clothing]')

    h_val = str(ws.cell(row=row_idx, column=8).value or '').strip()
    if h_val != '材质|棉质': issues.append('Row' + str(row_idx) + ' H=' + repr(h_val) + ' [标准=材质|棉质]')

    m_val = str(ws.cell(row=row_idx, column=13).value or '').strip()
    if m_val != '件/个': issues.append('Row' + str(row_idx) + ' M=' + repr(m_val) + ' [标准=件/个]')

    o_val = str(ws.cell(row=row_idx, column=15).value or '').strip()
    if o_val != 'Y': issues.append('Row' + str(row_idx) + ' O=' + repr(o_val) + ' [标准=Y]')

    p_val = str(ws.cell(row=row_idx, column=16).value or '').strip()
    if p_val != '0.3': issues.append('Row' + str(row_idx) + ' P=' + repr(p_val) + ' [标准=0.3]')

    y_val = ws.cell(row=row_idx, column=25).value
    if not y_val or str(y_val).strip() != 'Size\nS\nM\nL\nXL':
        issues.append('Row' + str(row_idx) + ' Y=' + repr(str(y_val)[:30]) + ' [标准=Size\\nS\\nM\\nL\\nXL]')

    ad_val = ws.cell(row=row_idx, column=30).value
    if ad_val != 59: issues.append('Row' + str(row_idx) + ' AD=' + repr(ad_val) + ' [标准=59]')

    ae_val = ws.cell(row=row_idx, column=31).value
    if ae_val != 99: issues.append('Row' + str(row_idx) + ' AE=' + repr(ae_val) + ' [标准=99]')

    af_val = ws.cell(row=row_idx, column=32).value
    if af_val != 999: issues.append('Row' + str(row_idx) + ' AF=' + repr(af_val) + ' [标准=999]')

    for col_idx in [1, 3, 7, 14, 17, 18, 19, 23, 24, 26, 27, 29]:
        v_ = ws.cell(row=row_idx, column=col_idx).value
        if v_ is not None and str(v_).strip():
            issues.append('Row' + str(row_idx) + ' ' + col_letter(col_idx) + '=' + repr(str(v_)[:30]) + ' [标准=空]')

    e_val = str(ws.cell(row=row_idx, column=5).value or '')
    f_val = str(ws.cell(row=row_idx, column=6).value or '')
    e_count = 1 if e_val.startswith('http') else 0
    f_count = f_val.count('\n')+1 if f_val.startswith('http') else 0
    if e_count + f_count > 14:
        issues.append('Row' + str(row_idx) + ' 图片=' + str(e_count+f_count) + '>14')

if issues:
    print('  发现 ' + str(len(issues)) + ' 个问题:')
    for iss in issues[:20]:
        print('    ' + iss)
    if len(issues) > 20:
        print('    ... 还有 ' + str(len(issues)-20) + ' 个问题')
else:
    print('  [OK] 所有主行字段合规')

print()
print('=== SKU子行字段检查 ===')
sku_issues = []
sku_checked = 0
for row_idx in range(1, ws.max_row + 1):
    ab_val = str(ws.cell(row=row_idx, column=28).value or '').strip()
    if ab_val not in ['Size:M', 'Size:L', 'Size:XL']: continue

    ad = ws.cell(row=row_idx, column=30).value
    ae = ws.cell(row=row_idx, column=31).value
    af = ws.cell(row=row_idx, column=32).value

    if ad != 59: sku_issues.append('Row' + str(row_idx) + ' ' + ab_val + ' AD=' + repr(ad) + ' [标准=59]')
    if ae != 99: sku_issues.append('Row' + str(row_idx) + ' ' + ab_val + ' AE=' + repr(ae) + ' [标准=99]')
    if af != 999: sku_issues.append('Row' + str(row_idx) + ' ' + ab_val + ' AF=' + repr(af) + ' [标准=999]')

    for col_idx in [2, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 20, 21, 22, 24, 25]:
        v_ = ws.cell(row=row_idx, column=col_idx).value
        if v_ is not None and str(v_).strip():
            sku_issues.append('Row' + str(row_idx) + ' ' + ab_val + ' ' + col_letter(col_idx) + '有值=' + repr(str(v_)[:20]))

    sku_checked += 1

if sku_issues:
    print('  SKU子行发现 ' + str(len(sku_issues)) + ' 个问题:')
    for iss in sku_issues[:20]:
        print('    ' + iss)
else:
    print('  [OK] SKU子行字段合规 (AD=59, AE=99, AF=999, 其他字段为空)')

print()
print('=== D列Name字段 vs B列一致性 ===')
d_issues = []
for row_idx in range(1, ws.max_row + 1):
    ab_val = str(ws.cell(row=row_idx, column=28).value or '').strip()
    if ab_val != 'Size:S': continue
    b_val = str(ws.cell(row=row_idx, column=2).value or '').strip()
    d_val = str(ws.cell(row=row_idx, column=4).value or '')
    if not b_val or 'Name:' not in d_val: continue
    m = re.search(r'<a[^>]+>([^<]+)</a>', d_val)
    if m:
        d_brand = m.group(1).strip()
        if d_brand and len(d_brand) > 1 and d_brand not in b_val[:len(d_brand)+3]:
            d_issues.append('Row' + str(row_idx) + ' D品牌=' + repr(d_brand) + ' 不在B前缀 B=' + repr(b_val[:40]))
if d_issues:
    print('  发现 ' + str(len(d_issues)) + ' 个不一致:')
    for iss in d_issues[:10]: print('    ' + iss)
else:
    print('  [OK] D/B一致性通过')

print()
print('=== 标题唯一性检查 ===')
all_titles = []
for row_idx in range(1, ws.max_row + 1):
    ab_val = str(ws.cell(row=row_idx, column=28).value or '').strip()
    if ab_val == 'Size:S':
        b_val = str(ws.cell(row=row_idx, column=2).value or '').strip()
        if b_val: all_titles.append(b_val)

from collections import Counter
tc = Counter(all_titles)
dups = [(t, c) for t, c in tc.items() if c > 1]
print('  唯一主行标题=' + str(len(tc)) + ', 总主行=' + str(len(all_titles)))
if dups:
    print('  重复标题=' + str(len(dups)) + ':')
    for t, c in dups[:10]:
        print('    ' + t + ' x' + str(c))
else:
    print('  [OK] 无重复标题')

print()
print('=== 商品总数/品牌统计 ===')
main_rows = sum(1 for r in range(1, ws.max_row+1) if str(ws.cell(row=r, column=28).value or '').strip() == 'Size:S')
brands = []
for row_idx in range(1, ws.max_row + 1):
    if str(ws.cell(row=row_idx, column=28).value or '').strip() == 'Size:S':
        k_val = str(ws.cell(row=row_idx, column=11).value or '').strip()
        if k_val: brands.append(k_val)

bc = Counter(brands)
print('  主行(商品)总数=' + str(main_rows))
print('  品牌分布:')
for brand, cnt in bc.most_common():
    print('    ' + brand + '=' + str(cnt) + '个')

print()
print('=== W列(SEO URL Handle)应为空 ===')
w_non_empty = 0
for row_idx in range(1, ws.max_row + 1):
    if str(ws.cell(row=row_idx, column=28).value or '').strip() == 'Size:S':
        w_val = ws.cell(row=row_idx, column=23).value
        if w_val is not None and str(w_val).strip():
            w_non_empty += 1
print('  W列非空的主行=' + str(w_non_empty) + ' [标准=空]')

print()
print('=== X列(规格1)应为空 ===')
x_non_empty = 0
for row_idx in range(1, ws.max_row + 1):
    if str(ws.cell(row=row_idx, column=28).value or '').strip() == 'Size:S':
        x_val = ws.cell(row=row_idx, column=24).value
        if x_val is not None and str(x_val).strip():
            x_non_empty += 1
print('  X列非空的主行=' + str(x_non_empty) + ' [标准=空]')