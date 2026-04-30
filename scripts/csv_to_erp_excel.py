#!/usr/bin/env python3
"""
Convert CSV to ERP Excel import format (BAPE_0418 standard)

ERP标准：
- 售价 = 59，原价 = 99
- T/U/V = B列标题
- D列 Name:字段 = strip_brand(B列标题)
- 每商品4行：主行 + 3行SKU子行
- SKU子行只填 AB/AD/AE/AF
"""

import csv
import os
import re
import shutil
import uuid
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Column mapping (1-indexed)
COL_A = 1  # 商品ID (空)
COL_B = 2  # 商品标题*
COL_C = 3  # 副标题 (空)
COL_D = 4  # 商品描述
COL_E = 5  # 商品首图*
COL_F = 6  # 商品其他图片
COL_G = 7  # 关键信息 (空)
COL_H = 8  # 属性
COL_I = 9  # 商品上架* (N)
COL_J = 10  # 物流模板* (Clothing)
COL_K = 11  # 类别名称
COL_L = 12  # 标签
COL_M = 13  # 计量单位 (件/个)
COL_N = 14  # 商品备注 (空)
COL_O = 15  # 不记库存* (Y)
COL_P = 16  # 商品重量* (0.3)
COL_Q = 17  # 包装长度 (空)
COL_R = 18  # 包装宽度 (空)
COL_S = 19  # 包装高度 (空)
COL_T = 20  # SEO标题
COL_U = 21  # SEO描述
COL_V = 22  # SEO关键词
COL_W = 23  # SEO URL Handle (空)
COL_X = 24  # 规格1 (空)
COL_Y = 25  # 规格2
COL_Z = 26  # 规格3 (空)
COL_AA = 27  # 规格4 (空)
COL_AB = 28  # SKU值
COL_AC = 29  # SKU图片 (空)
COL_AD = 30  # 售价*
COL_AE = 31  # 原价
COL_AF = 32  # 库存
COL_AG = 33  # SKU (空)

# Standard values per BAPE_0418
DEFAULT_STATUS = "N"
DEFAULT_LOGISTICS = "Clothing"
DEFAULT_UNIT = "件/个"
DEFAULT_NO_STOCK = "Y"
DEFAULT_WEIGHT = 0.3
DEFAULT_PRICE = 59
DEFAULT_ORIG_PRICE = 99
DEFAULT_STOCK = 999
DEFAULT_SPEC2 = "Size\nS\nM\nL\nXL"

# Brand mapping
BRAND_MAP = {
    "louis vuitton": "Louis Vuitton",
    "louisvuitton": "Louis Vuitton",
    "balenciaga": "Balenciaga",
    "clot": "CLOT",
    "saint laurent": "Saint Laurent",
    "celine": "Celine",
    "prada": "Prada",
    "fendi": "Fendi",
    "dior": "Dior",
    "loewe": "Loewe",
    "dolce": "Dolce & Gabbana",
    "givenchy": "Givenchy",
    "burberry": "Burberry",
}

# Brands to strip from Name: field
BRAND_PREFIXES = [
    "BAPE",
    "Louis Vuitton",
    "LouisVuitton",
    "Balenciaga",
    "CLOT",
    "Saint Laurent",
    "Celine",
    "PRADA",
    "Prada",
    "Fendi",
    "DIOR",
    "Dior",
    "Loewe",
    "Givenchy",
    "Burberry",
    "Dolce & Gabbana",
    "Moncler",
    "Gucci",
    "Off-White",
    "OFF-WHITE",
]


def detect_brand(title: str) -> str:
    """Detect brand from product title"""
    t_lower = title.lower()
    for key, brand in BRAND_MAP.items():
        if key in t_lower:
            return brand
    # Fallback: first word
    first = title.split(" ", 1)[0].strip()
    return first if first else "T-Shirt"


def strip_brand(title: str) -> str:
    """Remove brand prefix from title for Name: field (keep rest of title)"""
    for brand in BRAND_PREFIXES:
        pattern = rf"^{re.escape(brand)}[/\s]+"
        if re.match(pattern, title, re.IGNORECASE):
            return re.sub(pattern, "", title, flags=re.IGNORECASE).strip()
    return title


def get_brand_slug(brand: str) -> str:
    """Get URL slug for brand"""
    slug_map = {
        "Louis Vuitton": "Louis-Vuitton",
        "Balenciaga": "Balenciaga",
        "CLOT": "CLOT",
        "Saint Laurent": "Saint-Laurent",
        "Celine": "Celine",
        "Prada": "Prada",
        "Fendi": "Fendi",
        "Dior": "Dior",
        "Loewe": "Loewe",
        "Dolce & Gabbana": "Dolce-Gabbana",
        "Givenchy": "Givenchy",
        "Burberry": "Burberry",
    }
    return slug_map.get(brand, brand.replace(" ", "-"))


def build_desc_html(title: str, brand: str) -> str:
    """Build description HTML per updated standard"""
    brand_slug = get_brand_slug(brand)
    name_without_brand = strip_brand(title)

    # Name field with brand link
    name_field = f'<span style="font-family: verdana, geneva, sans-serif;"><a href="https://www.stockxshoesvip.net/{brand_slug}/" rel="noopener" target="_blank">{brand}</a> {name_without_brand}</span>'

    # Category field - split into two links
    category_field = f'<span style="font-size: 14px;"><a href="https://www.stockxshoesvip.net/{brand_slug}-Clothes/" target="_self" class="third-link animation-underline">{brand}</a></span><a href="https://www.stockxshoesvip.net/{brand_slug}-Clothes/" target="_self"> Clothes</a>'

    return f"""<p><span style="font-family: Tahoma;"><span>Name: {name_field}</span></span></p>
<p>Category: {category_field}</p>
<p><span style="font-weight: bold;">More about&nbsp;</span><span style="font-weight: bold;"></span></p>
<p><a href="https://www.stockxshoesvip.net/Fear-Of-God-Clothing/" rel="noopener" target="_blank">Fear Of God Clothing</a>&nbsp; <a href="https://www.stockxshoesvip.net/Pants/" rel="noopener" target="_blank">Pants</a>&nbsp; <a href="https://www.stockxshoesvip.net/Denim-Tears/" rel="noopener" target="_blank">Denim Tears</a></p>
<p><b>Our Core Guarantees</b></p>
<ul>
<li><b>Exclusive <a href="https://www.stockxshoesvip.net/Stockxshoes-QC-Pics/" rel="noopener" target="_blank">QC Service</a>:</b> We provide free Quality Control (QC) pictures before shipment. You approve the exact item you will receive&mdash;if not satisfied, we offer free exchanges or refunds.</li>
<li><b>Premium Packaging:</b> All apparel comes with full brand packaging and original tags.</li>
<li><b>Worry-Free Logistics:</b> We handle secure delivery and customs clearance to ensure your package arrives safely.</li>
<li><b>100% Safe Shopping:</b> 30-day money-back guarantee with damage protection.</li>
</ul>
<p><b>Shipping &amp; Payment</b></p>
<ul>
<li><b>Delivery Time:</b> 7-18 Days (Minor 1-3 day delays are normal). Tracking number provided.</li>
<li><b>Shipping Methods:</b> FedEx / USPS / DHL / UPS / EMS / Royal Mail.</li>
<li><b>Payment Methods:</b> Credit/Debit Card, PayPal, Bank Transfer, Cash App, Zelle.</li>
</ul>
<p><b>About StockxShoesVIP</b></p>
<p>With 10 years of offline retail and 5 years of online excellence, we are your trusted source for <b>premium replica sneakers and streetwear</b>.</p>
<p><i>(Note: We are an independent supplier and not affiliated with the StockX platform. Please bookmark our official site: stockxshoesvip.net)</i></p>
<p><b>Contact Us</b></p>
<ul>
<li><b>WhatsApp/WeChat:</b> +86 189 5920 5893</li>
<li><b>Instagram:</b> @stockxshoesvip_com</li>
</ul>
<p><i>Buy with confidence, wear with confidence.</i></p>"""


def parse_csv(csv_path: str) -> list:
    """Parse CSV file and extract product data"""
    products = []

    with open(csv_path, "r", encoding="utf-8") as f:
        content = f.read()

    lines = content.strip().split("\n")

    i = 1  # Skip header
    while i < len(lines):
        line = lines[i]

        # Parse: No.,Title,FirstImg,"OtherImgs..."
        first_comma = line.find(",")
        if first_comma == -1:
            i += 1
            continue

        no = line[:first_comma].strip()
        rest = line[first_comma + 1 :]

        second_comma = rest.find(",")
        if second_comma == -1:
            i += 1
            continue

        title = rest[:second_comma].strip()
        rest2 = rest[second_comma + 1 :]

        third_comma = rest2.find(",")
        if third_comma == -1:
            first_img = rest2.strip().strip('"')
            other_imgs = ""
        else:
            first_img = rest2[:third_comma].strip().strip('"')
            other_imgs = rest2[third_comma + 1 :].strip()

        # Collect multi-line other images
        i += 1
        while i < len(lines) and not lines[i][0].isdigit():
            other_imgs += "\n" + lines[i].strip().strip('"')
            i += 1

        # Parse and dedupe images
        img_list = []
        seen = set()
        for img in other_imgs.split("\n"):
            img = img.strip().strip('"')
            if img and img not in seen:
                # Normalize URL
                img = img.replace("https://", "http://")
                img = img.replace("photo.yupoo.com", "pic.yupoo.com")
                if img not in seen:
                    img_list.append(img)
                    seen.add(img)

        # Limit to 13 other images (total 14 with first image)
        img_list = img_list[:13]

        products.append(
            {
                "no": int(no) if no.isdigit() else len(products) + 1,
                "title": title,
                "first_img": first_img,
                "other_imgs": img_list,
                "brand": detect_brand(title),
            }
        )

    return products


def create_erp_excel(products: list, output_path: str):
    """Create ERP Excel file from product data"""
    tmp_dir = f".dumate/xlsx-{uuid.uuid4()}"
    os.makedirs(tmp_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息"

    # Header style
    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers (row 1) - 33 columns
    headers = {
        COL_A: "商品ID",
        COL_B: "商品标题*",
        COL_C: "副标题",
        COL_D: "商品描述",
        COL_E: "商品首图*",
        COL_F: "商品其他图片",
        COL_G: "关键信息",
        COL_H: "属性",
        COL_I: "商品上架*",
        COL_J: "物流模板*",
        COL_K: "类别名称",
        COL_L: "标签",
        COL_M: "计量单位",
        COL_N: "商品备注",
        COL_O: "不记库存*",
        COL_P: "商品重量*",
        COL_Q: "包装长度",
        COL_R: "包装宽度",
        COL_S: "包装高度",
        COL_T: "SEO标题",
        COL_U: "SEO描述",
        COL_V: "SEO关键词",
        COL_W: "SEO URL Handle",
        COL_X: "规格1",
        COL_Y: "规格2",
        COL_Z: "规格3",
        COL_AA: "规格4",
        COL_AB: "SKU值",
        COL_AC: "SKU图片",
        COL_AD: "售价*",
        COL_AE: "原价",
        COL_AF: "库存",
        COL_AG: "SKU",
    }

    for col, header in headers.items():
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hf
        cell.fill = hfill
        cell.border = border

    # Set column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 80
    ws.column_dimensions["I"].width = 8
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 40
    ws.column_dimensions["M"].width = 10
    ws.column_dimensions["P"].width = 10
    ws.column_dimensions["T"].width = 40
    ws.column_dimensions["U"].width = 40
    ws.column_dimensions["V"].width = 40
    ws.column_dimensions["Y"].width = 15
    ws.column_dimensions["AB"].width = 10
    ws.column_dimensions["AD"].width = 8
    ws.column_dimensions["AE"].width = 8
    ws.column_dimensions["AF"].width = 8

    # Data rows
    row = 2
    seen_titles = {}

    for p in products:
        # Deduplicate title
        title = p["title"]
        if title in seen_titles:
            seen_titles[title] += 1
            suffix = f" {seen_titles[title]}"
            if len(title) + len(suffix) <= 255:
                title = title + suffix
            else:
                title = title[: 255 - len(suffix)] + suffix
        else:
            seen_titles[title] = 1

        brand = p["brand"]
        desc = build_desc_html(title, brand)
        other_imgs_str = "\n".join(p["other_imgs"][:13])

        # Create 4 rows per product (Size: S/M/L/XL)
        sizes = ["Size:S", "Size:M", "Size:L", "Size:XL"]

        for idx, size in enumerate(sizes):
            if idx == 0:
                # Main row - fill all fields
                ws.cell(row=row, column=COL_A, value="").border = border  # 商品ID (空)
                ws.cell(row=row, column=COL_B, value=title).border = border
                ws.cell(row=row, column=COL_C, value="").border = border  # 副标题 (空)
                ws.cell(row=row, column=COL_D, value=desc).border = border
                ws.cell(row=row, column=COL_E, value=p["first_img"]).border = border
                ws.cell(row=row, column=COL_F, value=other_imgs_str).border = border
                ws.cell(
                    row=row, column=COL_G, value=""
                ).border = border  # 关键信息 (空)
                ws.cell(
                    row=row, column=COL_H, value="材质|棉质"
                ).border = border  # 属性
                ws.cell(
                    row=row, column=COL_I, value=DEFAULT_STATUS
                ).border = border  # N
                ws.cell(
                    row=row, column=COL_J, value=DEFAULT_LOGISTICS
                ).border = border  # Clothing
                ws.cell(row=row, column=COL_K, value=brand).border = border
                ws.cell(
                    row=row, column=COL_L, value=title
                ).border = border  # 标签 = 标题
                ws.cell(
                    row=row, column=COL_M, value=DEFAULT_UNIT
                ).border = border  # 件/个
                ws.cell(
                    row=row, column=COL_N, value=""
                ).border = border  # 商品备注 (空)
                ws.cell(
                    row=row, column=COL_O, value=DEFAULT_NO_STOCK
                ).border = border  # Y
                ws.cell(
                    row=row, column=COL_P, value=DEFAULT_WEIGHT
                ).border = border  # 0.3
                ws.cell(
                    row=row, column=COL_Q, value=""
                ).border = border  # 包装长度 (空)
                ws.cell(
                    row=row, column=COL_R, value=""
                ).border = border  # 包装宽度 (空)
                ws.cell(
                    row=row, column=COL_S, value=""
                ).border = border  # 包装高度 (空)
                # T/U/V per TheNorthface_ERP_导入模板_0418.xlsx standard
                seo_title = f"Stockx Replica Streetwear | Top Quality 1:1 {title} - stockxshoesvip.net"
                seo_desc = f"Buy Best 1:1 Replica Clothing on Stockxshoesvip.net. Perfect {title}. 100% safe shipping, free QC confirmation, and easy returns."
                seo_keyword = title
                ws.cell(
                    row=row, column=COL_T, value=seo_title
                ).border = border  # SEO标题
                ws.cell(
                    row=row, column=COL_U, value=seo_desc
                ).border = border  # SEO描述
                ws.cell(
                    row=row, column=COL_V, value=seo_keyword
                ).border = border  # SEO关键词 = 标题
                ws.cell(
                    row=row, column=COL_W, value=""
                ).border = border  # SEO URL Handle (空)
                ws.cell(row=row, column=COL_X, value="").border = border  # 规格1 (空)
                ws.cell(
                    row=row, column=COL_Y, value=DEFAULT_SPEC2
                ).border = border  # Size\nS\nM\nL\nXL
                ws.cell(row=row, column=COL_Z, value="").border = border  # 规格3 (空)
                ws.cell(row=row, column=COL_AA, value="").border = border  # 规格4 (空)
                ws.cell(row=row, column=COL_AB, value=size).border = border  # Size:S
                ws.cell(
                    row=row, column=COL_AC, value=""
                ).border = border  # SKU图片 (空)
                ws.cell(
                    row=row, column=COL_AD, value=DEFAULT_PRICE
                ).border = border  # 59
                ws.cell(
                    row=row, column=COL_AE, value=DEFAULT_ORIG_PRICE
                ).border = border  # 99
                ws.cell(
                    row=row, column=COL_AF, value=DEFAULT_STOCK
                ).border = border  # 999
                ws.cell(row=row, column=COL_AG, value="").border = border  # SKU (空)
            else:
                # SKU sub-row - only fill AB/AD/AE/AF
                ws.cell(row=row, column=COL_AB, value=size).border = border
                ws.cell(row=row, column=COL_AD, value=DEFAULT_PRICE).border = border
                ws.cell(
                    row=row, column=COL_AE, value=DEFAULT_ORIG_PRICE
                ).border = border
                ws.cell(row=row, column=COL_AF, value=DEFAULT_STOCK).border = border

            row += 1

    # Save to temp file
    tmp_file = os.path.join(tmp_dir, "erp_import.xlsx")
    wb.save(tmp_file)

    # Copy to output
    shutil.copy2(tmp_file, output_path)

    # Cleanup
    shutil.rmtree(tmp_dir, ignore_errors=True)

    return len(products)


def main():
    import sys

    if len(sys.argv) < 2:
        print("Usage: python csv_to_erp_excel.py <input.csv> [output.xlsx]")
        return 1

    csv_path = sys.argv[1]

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base = Path(csv_path).stem
        output_path = f"inputs/{base}_ERP.xlsx"

    print(f"Parsing CSV: {csv_path}")
    products = parse_csv(csv_path)
    print(f"Found {len(products)} products")

    print(f"Creating ERP Excel: {output_path}")
    count = create_erp_excel(products, output_path)

    print(
        f"\n[OK] Done! Created {count} products (4 rows each = {count * 4} total rows)"
    )
    print(f"Output: {output_path}")
    print(f"\nERP Standard:")
    print(f"  - Price = {DEFAULT_PRICE}")
    print(f"  - Compare-at = {DEFAULT_ORIG_PRICE}")
    print(f"  - T/U/V = B column title")
    print(f"  - I column = N (offline)")

    return 0


if __name__ == "__main__":
    exit(main())
