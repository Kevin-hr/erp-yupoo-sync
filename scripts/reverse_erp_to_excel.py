# -*- coding: utf-8 -*-
"""
逆向流程演示脚本 - 从ERP提取商品数据填入Excel模板
Reverse Flow Demo - Extract product from ERP and fill into Excel template
"""

import asyncio
import json
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os


async def extract_product_from_erp(product_id_or_name=None):
    """从ERP提取商品数据"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()

        with open("logs/cookies.json", "r") as f:
            cookies = json.load(f)
        await context.add_cookies(cookies)

        page = await context.new_page()

        # Navigate to product list
        await page.goto("https://www.mrshopplus.com/#/product/list_DTB_proProduct")
        await page.wait_for_load_state("networkidle", timeout=30000)
        await asyncio.sleep(8)

        # Get product data from first row (demo)
        rows = await page.query_selector_all(".el-table__row")
        if not rows:
            return None

        # Click first product to edit
        first_row = rows[0]
        clickables = await first_row.query_selector_all(
            'a, .cell a, span[class*="name"]'
        )
        for cl in clickables:
            txt = await cl.inner_text()
            if txt and len(txt.strip()) > 3:
                await cl.click()
                await asyncio.sleep(5)
                break

        # Extract form data
        form_data = await page.evaluate("""() => {
            const data = {};
            
            // Get inputs
            const inputs = document.querySelectorAll('input');
            inputs.forEach((inp, i) => {
                const name = inp.name || inp.id || `input_${i}`;
                const type = inp.type;
                if (type === 'text' || type === 'number') {
                    data[name] = inp.value;
                }
            });
            
            // Get TinyMCE content
            const mce_iframes = document.querySelectorAll('iframe[id^="vue-tinymce"]');
            const mce_contents = [];
            mce_iframes.forEach((iframe) => {
                try {
                    const doc = iframe.contentDocument || iframe.contentWindow.document;
                    mce_contents.push(doc.body.innerText);
                } catch(e) {
                    mce_contents.push('');
                }
            });
            data['mce_contents'] = mce_contents;
            
            // Get images (thumbnail URLs)
            const images = [];
            const img_elements = document.querySelectorAll('.el-image-viewer__wrapper img, .upload-container img, img[src*="image"]');
            img_elements.forEach(img => {
                if (img.src && !img.src.includes('logo')) {
                    images.push(img.src);
                }
            });
            data['images'] = images;
            
            return data;
        }""")

        return form_data


def fill_excel_template(form_data, output_path, test_product_name=None):
    """将数据填入Excel模板"""
    wb = load_workbook("商品导入模板.xlsx")
    ws = wb.active

    # 解析提取的数据
    product_name = test_product_name or form_data.get("input_1", "")
    description = (
        form_data.get("mce_contents", [""])[0] if form_data.get("mce_contents") else ""
    )
    images = form_data.get("images", [])

    # 获取价格数据 - 多个SKU价格
    prices = []
    for key in ["input_25", "input_31", "input_37", "input_43", "input_49"]:
        if form_data.get(key):
            try:
                prices.append(float(form_data.get(key, 0)))
            except:
                pass

    original_prices = []
    for key in ["input_26", "input_32", "input_38", "input_44", "input_50"]:
        if form_data.get(key):
            try:
                original_prices.append(float(form_data.get(key, 0)))
            except:
                pass

    # 准备行数据 - 参照模板第4行格式
    row_data = [
        "",  # 商品ID (系统ID，空白表示新增)
        product_name,  # 商品名称
        "",  # 规格
        "",  # 商品描述
        images[0] if images else "",  # 商品主图URL
        "\n".join(images[:14]) if images else "",  # 商品详情图片URL
        "",  # 规格信息
        "",  # 分类
        "Y",  # 上架
        "默认模板",  # 模板
        "",  # 系列
        "",  # 标签
        "",  # 库存位置
        "",  # 商品备注
        "Y",  # 必填
        "0.5",  # 重量kg
        "",  # 包装长cm
        "",  # 包装宽cm
        "",  # 包装高cm
        "",  # SEO描述
        "",  # SEO标题
        "",  # SEO关键词
        "",  # SEO URL Handle
    ]

    # 找到第一个空行 (Row 4 based on template structure)
    # Template has: Row 1-3 = headers/instructions, Row 4+ = data
    target_row = 4

    # 填充主数据行
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=target_row, column=col_idx, value=value)

    # 填充SKU数据 (Columns 24-33)
    # SKU值格式: "Color:Yellow\\nSize:41"
    sku_row_start = 24

    # 如果有测试产品名称，创建测试SKU
    if test_product_name:
        # 测试用例SKU
        sku_value = f"Color:Black\\nSize:S/M/L/XL/XXL"
        ws.cell(row=target_row, column=28, value=sku_value)  # SKU值
        ws.cell(row=target_row, column=29, value=images[0] if images else "")  # SKU图片
        ws.cell(row=target_row, column=30, value=88.99)  # 售价
        ws.cell(row=target_row, column=31, value=149.99)  # 原价
        ws.cell(row=target_row, column=32, value=100)  # 数量
        ws.cell(row=target_row, column=33, value="DESC-POLO-BLK-S")  # SKU
    else:
        # 使用提取的ERP数据
        for i, (price, orig_price) in enumerate(zip(prices[:5], original_prices[:5])):
            ws.cell(row=target_row, column=28 + i * 5, value=f"SKU_{i + 1}")
            ws.cell(
                row=target_row,
                column=29 + i * 5,
                value=images[i] if i < len(images) else "",
            )
            ws.cell(
                row=target_row, column=30 + i * 5, value=price if price > 0 else 88.99
            )
            ws.cell(
                row=target_row,
                column=31 + i * 5,
                value=orig_price if orig_price > 0 else 149.99,
            )
            ws.cell(row=target_row, column=32 + i * 5, value=100)

    wb.save(output_path)
    print(f"Excel saved to: {output_path}")
    return output_path


async def main():
    print("=== ERP逆向提取演示 ===")
    print("目标: DescenteTraining Short Sleeve Polo Black")
    print()

    # 提取ERP商品数据 (使用BAPE产品作为演示)
    print("1. 从ERP提取商品数据...")
    form_data = await extract_product_from_erp()

    if form_data:
        print(f"   提取到数据:")
        print(f"   - 商品名称: {form_data.get('input_1', 'N/A')}")
        print(f"   - 图片数量: {len(form_data.get('images', []))}")
        print(f"   - 描述长度: {len(str(form_data.get('mce_contents', [''])[0]))}")

    # 填入Excel模板
    print()
    print("2. 填入Excel模板...")
    output_dir = "logs"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "商品导入模板_已填充.xlsx")

    # 使用测试用例产品名称填充
    fill_excel_template(
        form_data,
        output_path,
        test_product_name="DescenteTraining Short Sleeve Polo Black",
    )

    print()
    print("3. 完成!")
    print(f"输出文件: {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
