# -*- coding: utf-8 -*-
"""
Sync all fields related to column B (商品标题) per ERP_EXCEL_STANDARD_BAPE0418.md v2.0.0

Fields to sync:
- L (标签) = B列标题
- V (SEO关键词) = B列标题
- T (SEO标题) = Stockx Replica Streetwear | Top Quality 1:1 {标题} - stockxshoesvip.net
- U (SEO描述) = Buy Best 1:1 Replica Clothing on Stockxshoesvip.net. Perfect {标题}...
- D (商品描述) = HTML模板，Name字段使用去品牌后的标题
- K (类别名称) = 品牌名
"""

import re
from pathlib import Path
from openpyxl import load_workbook

# Brand URL slug mapping
BRAND_SLUGS = {
    "Louis Vuitton": "Louis-Vuitton",
    "Balenciaga": "Balenciaga",
    "Burberry": "Burberry",
    "Dior": "Dior",
    "Gucci": "Gucci",
    "GUCCI": "Gucci",
    "Loewe": "Loewe",
    "LOEWE": "Loewe",
    "Moncler": "Moncler",
    "MONCLER": "Moncler",
    "MIU MIU": "MIU-MIU",
    "Miu Miu": "MIU-MIU",
    "Givenchy": "Givenchy",
    "GIVENCHY": "Givenchy",
    "Celine": "Celine",
    "CELINE": "Celine",
    "PRADA": "Prada",
    "Prada": "Prada",
    "Off-White": "Off-White",
    "Alexander Wang": "Alexander-Wang",
    "Ami Paris": "Ami-Paris",
    "AMI Paris": "Ami-Paris",
    "Thom Browne": "Thom-Browne",
}

# HTML template for D column
HTML_TEMPLATE = """<p><span style="font-family: Tahoma;"><span>Name: <span style="font-family: verdana, geneva, sans-serif;"><a href="https://www.stockxshoesvip.net/{brand_slug}/" rel="noopener" target="_blank">{brand}</a> {name_without_brand}</span></span></span></p>
<p>Category: <span style="font-size: 14px;"><a href="https://www.stockxshoesvip.net/{brand_slug}-Clothes/" target="_self" class="third-link animation-underline">{brand}</a></span><a href="https://www.stockxshoesvip.net/{brand_slug}-Clothes/" target="_self"> Clothes</a></p>
<p><span style="font-weight: bold;">More about&nbsp;</span><span style="font-weight: bold;"></span></p>
<p><a href="https://www.stockxshoesvip.net/Fear-Of-God-Clothing/" rel="noopener" target="_blank">Fear Of God Clothing</a>&nbsp; <a href="https://www.stockxshoesvip.net/Pants/" rel="noopener" target="_blank">Pants</a>&nbsp; <a href="https://www.stockxshoesvip.net/Denim-Tears/" rel="noopener" target="_blank">Denim Tears</a></p>
<p><b>Our Core Guarantees</b></p>
<ul>
<li><b>Exclusive <a href="https://www.stockxshoesvip.net/Stockxshoes-QC-Pics/" rel="noopener" target="_blank">QC Service</a>:</b> We provide free Quality Control (QC) pictures before shipment. You approve the exact item you will receive&mdash;if not satisfied, we offer free exchanges or refunds.</li>
<li><b>Premium Packaging:</b> All apparel comes with full brand packaging and original tags.</li>
<li><b>Worry-Free Logistics:</b> We handle secure delivery and customs clearance to ensure your package arrives safely.</li>
<li><b>100% Safe Shopping:</b> 30-day money-back guarantee with damage protection.</li>
</ul>
<p><b>Shipping &amp; Payment</b></p>
<ul>
<li><b>Delivery Time:</b> 7-18 Days (Minor 1-3 day delays are normal). Tracking number provided.</li>
<li><b>Shipping Methods:</b> FedEx / USPS / DHL / UPS / EMS / Royal Mail.</li>
<li><b>Payment Methods:</b> Credit/Debit Card, PayPal, Bank Transfer, Cash App, Zelle.</li>
</ul>
<p><b>About StockxShoesVIP</b></p>
<p>With 10 years of offline retail and 5 years of online excellence, we are your trusted source for <b>premium replica sneakers and streetwear</b>.</p>
<p><i>(Note: We are an independent supplier and not affiliated with the StockX platform. Please bookmark our official site: stockxshoesvip.net)</i></p>
<p><b>Contact Us</b></p>
<ul>
<li><b>WhatsApp/WeChat:</b> +86 189 5920 5893</li>
<li><b>Instagram:</b> @stockxshoesvip_com</li>
</ul>
<p><i>Buy with confidence, wear with confidence.</i></p>"""


def extract_brand(title):
    """Extract brand name from title."""
    if not title:
        return "Unknown", "Unknown"

    # Try to match brand at the start of title
    for brand in sorted(BRAND_SLUGS.keys(), key=len, reverse=True):
        if title.startswith(brand):
            slug = BRAND_SLUGS.get(brand, brand.replace(" ", "-"))
            return brand, slug

    # Default: first word as brand
    first_word = title.split()[0] if title.split() else "Unknown"
    return first_word, first_word.replace(" ", "-")


def strip_brand(title, brand):
    """Remove brand prefix from title for Name field."""
    if not title or not brand:
        return title or ""

    # Remove brand from start of title
    pattern = re.compile(r"^" + re.escape(brand) + r"\s*", re.IGNORECASE)
    result = pattern.sub("", title)
    return result.strip() if result.strip() else title


def build_seo_title(title):
    """Build SEO title: Stockx Replica Streetwear | Top Quality 1:1 {标题} - stockxshoesvip.net"""
    return f"Stockx Replica Streetwear | Top Quality 1:1 {title} - stockxshoesvip.net"


def build_seo_description(title):
    """Build SEO description."""
    return f"Buy Best 1:1 Replica Clothing on Stockxshoesvip.net. Perfect {title}. 100% safe shipping, free QC confirmation, and easy returns."


def build_description_html(title):
    """Build D column HTML with Name field."""
    brand, slug = extract_brand(title)
    name_without_brand = strip_brand(title, brand)

    return HTML_TEMPLATE.format(
        brand_slug=slug, brand=brand, name_without_brand=name_without_brand
    )


def main():
    # Load Excel
    excel_path = Path("inputs/ERP_Export_2026-04-28_96products_final.xlsx")
    wb = load_workbook(excel_path)

    # Get sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "商品信息" in name:
            sheet_name = name
            break
    sheet = wb[sheet_name]

    print("=== Syncing B-related fields ===\n")

    # Process each main row (every 4th row starting from 4)
    updated_count = 0
    for row in range(4, sheet.max_row + 1, 4):
        title = sheet[f"B{row}"].value
        if not title:
            continue

        # Extract brand
        brand, slug = extract_brand(title)

        # L (标签) = B列标题
        sheet[f"L{row}"] = title

        # V (SEO关键词) = B列标题
        sheet[f"V{row}"] = title

        # T (SEO标题)
        seo_title = build_seo_title(title)
        sheet[f"T{row}"] = seo_title

        # U (SEO描述)
        seo_desc = build_seo_description(title)
        sheet[f"U{row}"] = seo_desc

        # D (商品描述) - HTML with Name field
        html_desc = build_description_html(title)
        sheet[f"D{row}"] = html_desc

        # K (类别名称) = 品牌名
        sheet[f"K{row}"] = brand

        updated_count += 1

    # Save
    wb.save(excel_path)
    print(f"Updated {updated_count} products")
    print(f"Saved to: {excel_path}")

    # Verification
    print("\n=== Verification ===")
    for row in [4, 8, 32, 36]:
        b = sheet[f"B{row}"].value
        l = sheet[f"L{row}"].value
        k = sheet[f"K{row}"].value
        t = sheet[f"T{row}"].value
        v = sheet[f"V{row}"].value

        print(f"\nRow {row}:")
        print(f"  B (Title): {b}")
        print(f"  K (Category): {k}")
        print(f"  L (Tag): {l[:50]}..." if l and len(l) > 50 else f"  L (Tag): {l}")
        print(
            f"  V (SEO Keywords): {v[:50]}..."
            if v and len(v) > 50
            else f"  V (SEO Keywords): {v}"
        )
        print(
            f"  T (SEO Title): {t[:60]}..."
            if t and len(t) > 60
            else f"  T (SEO Title): {t}"
        )


if __name__ == "__main__":
    main()
