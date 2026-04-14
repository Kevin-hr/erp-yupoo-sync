#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP 抓取 v2 - 直接用 Playwright API 操作搜索，JS evaluate 兜底
"""

import asyncio
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import async_playwright

ROOT_DIR = Path(__file__).parent.parent
COOKIES_FILE = ROOT_DIR / "logs" / "cookies.json"
TEMPLATE_FILE = ROOT_DIR / "商品导入模板.xlsx"
OUTPUT_DIR = ROOT_DIR / "output"
SCREENSHOT_DIR = ROOT_DIR / "screenshots"
LOG_DIR = ROOT_DIR / "logs"
for d in [OUTPUT_DIR, LOG_DIR, SCREENSHOT_DIR]: d.mkdir(exist_ok=True)

logger = logging.getLogger("erp_v2")
logger.setLevel(logging.INFO)
fh = logging.FileHandler(LOG_DIR / f"erp_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log", encoding='utf-8')
logger.addHandler(fh)
sh = logging.StreamHandler(); logger.addHandler(sh)

ERP_BASE = "https://www.mrshopplus.com"
PRODUCT_LIST_URL = f"{ERP_BASE}/#/product/list_DTB_proProduct"


class ERPProduct:
    def __init__(self):
        self.product_id = ""; self.title = ""; self.subtitle = ""
        self.description = ""; self.primary_image = ""; self.other_images = ""
        self.key_info = ""; self.attributes = ""; self.listed = "N"
        self.shipping_template = "默认模板"; self.category = ""
        self.tags = ""; self.unit = "件/个"; self.notes = ""
        self.no_inventory = "Y"; self.weight = 0.5
        self.pack_length = 0.0; self.pack_width = 0.0; self.pack_height = 0.0
        self.seo_title = ""; self.seo_description = ""; self.seo_keywords = ""
        self.seo_handle = ""; self.spec1_name = ""; self.spec2_name = ""
        self.spec3_name = ""; self.spec4_name = ""; self.sku_values = ""
        self.sku_images = ""; self.price = 99.0; self.original_price = 199.0
        self.stock = 100; self.sku_code = ""

    def to_row(self):
        return [
            self.product_id, self.title, self.subtitle, self.description,
            self.primary_image, self.other_images, self.key_info, self.attributes,
            self.listed, self.shipping_template, self.category, self.tags, self.unit,
            self.notes, self.no_inventory, self.weight, self.pack_length,
            self.pack_width, self.pack_height, self.seo_title, self.seo_description,
            self.seo_keywords, self.seo_handle, self.spec1_name, self.spec2_name,
            self.spec3_name, self.spec4_name, self.sku_values, self.sku_images,
            self.price, self.original_price, self.stock, self.sku_code
        ]


def fill_excel(product, output_path):
    shutil.copy(TEMPLATE_FILE, output_path)
    wb = load_workbook(output_path)
    ws = wb["商品信息"]
    row = 3
    for r in range(1, 5):
        val = ws.cell(row=r, column=2).value
        if val and "商品标题" in str(val):
            row = r + 1; break
    for col_idx, value in enumerate(product.to_row(), start=1):
        ws.cell(row=row, column=col_idx, value=value)
    wb.save(output_path)


async def get_table_rows_via_js(page) -> list:
    """通过 JS 获取表格所有行"""
    data = await page.evaluate("""
        () => {
            const table = document.querySelector('.el-table__body-wrapper');
            if (!table) return { error: 'no table', rows: [] };
            const trs = table.querySelectorAll('tr.el-table__row');
            const rows = [];
            trs.forEach((tr, i) => {
                const cells = tr.querySelectorAll('td');
                const cellData = [];
                cells.forEach((cell, ci) => {
                    const imgs = [];
                    cell.querySelectorAll('img').forEach(img => {
                        if (img.src && !img.src.includes('logo') && !img.src.includes('head-logo')) {
                            imgs.push(img.src);
                        }
                    });
                    cellData.push({
                        col: ci,
                        text: cell.innerText.trim().substring(0, 80),
                        imgs: imgs
                    });
                });
                rows.push({ index: i, cells: cellData });
            });
            return { rows, count: rows.length };
        }
    """)
    return data


async def main(keyword: str):
    logger.info(f"=== ERP Extractor v2: {keyword} ===")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={"width": 1920, "height": 1080})

        if COOKIES_FILE.exists():
            cookies = json.load(open(COOKIES_FILE, 'r', encoding='utf-8'))
            await context.add_cookies(cookies)
            logger.info(f"[Init] Loaded {len(cookies)} cookies")

        page = await context.new_page()

        # ===== 步骤1: 导航到商品列表 =====
        await page.goto(PRODUCT_LIST_URL, timeout=30000)
        await asyncio.sleep(5)

        # ===== 步骤2: 探索页面结构 =====
        info = await page.evaluate("""
            () => {
                const inputs = [];
                document.querySelectorAll('.el-input__inner').forEach(el => {
                    if (el.offsetParent !== null) {
                        inputs.push({ placeholder: el.placeholder, value: el.value, class: el.className.substring(0, 60) });
                    }
                });
                return { inputs };
            }
        """)
        logger.info(f"[Explore] Inputs: {[i['placeholder'] for i in info['inputs']]}")

        # ===== 步骤3: 填充搜索框 - 用 Playwright 原生 API =====
        # 搜索框placeholder是"请输入商品名称/编码"
        search_filled = False
        for inp in info['inputs']:
            if '商品名称' in inp['placeholder'] or '编码' in inp['placeholder']:
                try:
                    # 用 Playwright locator 精确定位
                    locator = page.locator(f".el-input__inner[placeholder='{inp['placeholder']}']")
                    await locator.clear()
                    await locator.fill(keyword)
                    logger.info(f"[Search] Filled: {inp['placeholder']} = {keyword}")
                    search_filled = True
                    break
                except Exception as e:
                    logger.warning(f"[Search] Fill error: {e}")

        if not search_filled:
            # 兜底：直接用 page.locator 找搜索框
            try:
                loc = page.locator(".el-input__inner").filter(placeholder="请输入商品名称/编码")
                await loc.clear()
                await loc.fill(keyword)
                search_filled = True
                logger.info("[Search] Filled via filter locator")
            except Exception as e:
                logger.warning(f"[Search] Filter locator failed: {e}")

        await asyncio.sleep(2)

        # ===== 步骤4: 点击查询按钮 =====
        try:
            # 找"查询"按钮
            btns = await page.query_selector_all("button")
            for btn in btns:
                txt = await btn.inner_text()
                if '查询' in txt:
                    await btn.click()
                    logger.info(f"[Search] Clicked button: {txt}")
                    break
            else:
                # 尝试用 Playwright API
                await page.get_by_role("button", name="查询").click()
                logger.info("[Search] Clicked via get_by_role")
        except Exception as e:
            logger.warning(f"[Search] Button click: {e}")
            # 兜底：按回车
            await page.keyboard.press("Enter")
            logger.info("[Search] Pressed Enter")

        await asyncio.sleep(4)  # 等待搜索结果加载
        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_v2_search.png"), timeout=10000)

        # ===== 步骤5: 获取表格数据 =====
        table_data = await get_table_rows_via_js(page)
        logger.info(f"[Table] Found {table_data.get('count', 0)} rows")
        for row in table_data.get('rows', []):
            logger.info(f"[Table] Row {row['index']}: {[c['text'] for c in row['cells']]}")

        # ===== 步骤6: 找匹配行 =====
        matched_row = None
        for row in table_data.get('rows', []):
            row_text = ' '.join([c['text'] for c in row['cells']])
            if keyword.lower() in row_text.lower():
                matched_row = row
                logger.info(f"[Match] Row {row['index']}: {row_text[:100]}")
                break

        # ===== 步骤7: 构建商品数据 =====
        product = ERPProduct()
        product.title = keyword

        if matched_row:
            for cell in matched_row['cells']:
                col = cell['col']
                text = cell['text']
                imgs = cell.get('imgs', [])

                # 根据列位置判断含义（从截图已知顺序）
                # col0=空, col1=商品名称, col2=时间, col3=售价比, col4=日期, col5=原价?, col6=库存?, col7=?col8=?
                if col == 1:
                    product.title = text
                elif col == 2:
                    pass  # 时间
                elif col == 3:
                    # 价格（从截图看是售价）
                    try:
                        nums = re.findall(r'[\d.]+', text)
                        if nums:
                            product.price = float(nums[0])
                    except: pass
                elif col == 4:
                    pass  # 日期
                elif col == 5:
                    try:
                        nums = re.findall(r'[\d.]+', text)
                        if nums: product.original_price = float(nums[0])
                    except: pass
                elif col == 6:
                    try:
                        product.stock = int(re.findall(r'\d+', text)[0])
                    except: pass
                elif col == 8:
                    product.listed = 'Y' if '上架' in text or text == 'Y' else 'N'

                # 图片
                for img_url in imgs:
                    if img_url and not product.primary_image:
                        product.primary_image = img_url
                    elif img_url and product.primary_image:
                        if product.other_images:
                            product.other_images += '\n' + img_url
                        else:
                            product.other_images = img_url

        # 尝试从JS获取更多行信息（如有价格列）
        all_rows_data = await page.evaluate("""
            () => {
                const table = document.querySelector('.el-table__body-wrapper');
                if (!table) return [];
                const rows = table.querySelectorAll('tr.el-table__row');
                const results = [];
                rows.forEach((tr, i) => {
                    const cells = tr.querySelectorAll('td');
                    if (cells.length >= 9) {
                        results.push({
                            col0: cells[0].innerText.trim(),
                            col1: cells[1].innerText.trim().substring(0,50),
                            col2: cells[2].innerText.trim(),
                            col3: cells[3].innerText.trim(),  // 售价比(折)
                            col4: cells[4].innerText.trim(),  // 日期
                            col5: cells[5].innerText.trim(),  // 原价?
                            col6: cells[6].innerText.trim(),  // 库存?
                            col7: cells[7].innerText.trim(),
                            col8: cells[8].innerText.trim()   // 上架?
                        });
                    }
                });
                return results;
            }
        """)

        # 精确匹配目标商品行
        target_row = None
        for row in all_rows_data:
            if keyword.lower() in row['col1'].lower():
                target_row = row
                break

        if target_row:
            logger.info(f"[Target] MATCHED: {target_row}")
            product.title = target_row['col1']
            try:
                product.price = float(re.findall(r'[\d.]+', target_row['col3'])[0])
            except: pass
            try:
                product.original_price = float(re.findall(r'[\d.]+', target_row['col5'])[0])
            except: pass
            try:
                product.stock = int(re.findall(r'\d+', target_row['col6'])[0])
            except: pass
            product.listed = 'Y' if '上架' in target_row['col8'] or target_row['col8'] == 'Y' else 'N'

        # ===== 步骤8: 点击进入详情页 =====
        if target_row or matched_row:
            try:
                # 用JS点击商品名称单元格
                click_idx = matched_row['index'] if matched_row else 0
                await page.evaluate(f"""
                    () => {{
                        const table = document.querySelector('.el-table__body-wrapper');
                        if (!table) return;
                        const rows = table.querySelectorAll('tr.el-table__row');
                        if ({click_idx} < rows.length) {{
                            const cells = rows[{click_idx}].querySelectorAll('td');
                            // 点击商品名称（第2列，index 1）
                            if (cells[1]) cells[1].click();
                            else if (cells[0]) cells[0].click();
                        }}
                    }}
                """)
                logger.info(f"[Detail] Clicked row {click_idx}")
                await asyncio.sleep(5)
                await page.screenshot(path=str(SCREENSHOT_DIR / "erp_v2_detail.png"), timeout=10000)

                # 从详情页提取数据
                detail = await page.evaluate("""
                    () => {
                        const result = { url: window.location.href, price: null, originalPrice: null, imgs: [], desc: '' };

                        // 找价格输入框
                        const inputs = document.querySelectorAll('input');
                        inputs.forEach(inp => {
                            const ph = inp.placeholder || '';
                            const val = inp.value;
                            if (ph.includes('售价') && val) result.price = val;
                            if ((ph.includes('原价') || ph.includes('划线')) && val) result.originalPrice = val;
                        });

                        // 找所有图片
                        document.querySelectorAll('img').forEach(img => {
                            const src = img.src || '';
                            if (src && !src.includes('logo') && !src.includes('head-logo') && img.offsetParent !== null) {
                                result.imgs.push(src);
                            }
                        });

                        // TinyMCE iframe
                        const iframes = document.querySelectorAll('iframe');
                        for (const frame of iframes) {
                            try {
                                const body = frame.contentDocument.body;
                                if (body && body.innerText && body.innerText.trim()) {
                                    result.desc = body.innerText.trim().substring(0, 500);
                                    break;
                                }
                            } catch(e) {}
                        }

                        return result;
                    }
                """)
                logger.info(f"[Detail] URL: {detail['url']}")
                logger.info(f"[Detail] Price: {detail['price']}, Original: {detail['originalPrice']}")
                logger.info(f"[Detail] Desc: {detail['desc'][:80]}")
                logger.info(f"[Detail] Imgs: {detail['imgs'][:3]}")

                if detail.get('price'):
                    try: product.price = float(detail['price'])
                    except: pass
                if detail.get('originalPrice'):
                    try: product.original_price = float(detail['originalPrice'])
                    except: pass
                if detail.get('desc') and not product.description:
                    product.description = f"<p>{detail['desc']}</p>"
                if detail.get('imgs'):
                    for img in detail['imgs']:
                        if not product.primary_image:
                            product.primary_image = img
                        else:
                            product.other_images = (product.other_images + '\n' + img) if product.other_images else img
                            # 限制14张
                            existing = [product.primary_image] + product.other_images.split('\n')
                            if img not in existing:
                                all_imgs = existing + [img]
                                product.primary_image = all_imgs[0]
                                product.other_images = '\n'.join(all_imgs[1:15])

            except Exception as e:
                logger.warning(f"[Detail] Error: {e}")

        # SEO Handle
        product.seo_handle = re.sub(r'[^a-z0-9]+', '-', product.title.lower())
        product.seo_handle = re.sub(r'^-+|-+$', '', product.seo_handle)[:100]

        # ===== 步骤9: 写入 Excel =====
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = re.sub(r'[\\/:*?"<>|]', '', keyword[:20])
        output_path = OUTPUT_DIR / f"ERP提取_{safe}_{ts}.xlsx"
        fill_excel(product, output_path)

        logger.info(f"=== DONE ===")
        logger.info(f"Title: {product.title}")
        logger.info(f"Price: {product.price} / Original: {product.original_price}")
        logger.info(f"Stock: {product.stock}")
        logger.info(f"Listed: {product.listed}")
        logger.info(f"Primary Image: {product.primary_image[:80] if product.primary_image else 'NONE'}")
        logger.info(f"Other Images: {len(product.other_images.split(chr(10))) if product.other_images else 0} images")
        logger.info(f"Output: {output_path}")

        await asyncio.sleep(2)
        await browser.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("keyword", help="商品关键词")
    args = parser.parse_args()
    asyncio.run(main(args.keyword))
