#!/usr/bin/env python3
"""
Yupoo Product Extractor - Click on products to extract data
"""

import json
import os
import re
import shutil
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    print("Error: playwright not installed")
    sys.exit(1)

OUTPUT_DIR = Path("inputs")
AUTH_FILE = "yupoo_auth.json"

ENGLISH_NAMES = {
    "Louis Vuitton": "Louis Vuitton",
    "Burberry": "Burberry",
    "Loewe": "Loewe",
    "MIU MIU": "MIU MIU",
    "GUCCI": "Gucci",
    "Gucci": "Gucci",
    "Moncler": "Moncler",
    "Balenciaga": "Balenciaga",
    "Givenchy": "Givenchy",
    "Celine": "Celine",
    "Dior": "Dior",
    "OFF-WHITE": "OFF-WHITE",
    "Thom Browne": "Thom Browne",
    "PRADA": "Prada",
    "Alexander Wang": "Alexander Wang",
    "BAPE": "BAPE",
    "Bottega Veneta": "Bottega Veneta",
    "Maison Margiela": "Maison Margiela",
}


def generate_english_name(chinese_name: str, index: int) -> str:
    name = re.sub(r"^[hH]?\d+", "", chinese_name).strip()
    brand = None
    for k, v in ENGLISH_NAMES.items():
        if k in name or k in chinese_name:
            brand = v
            break

    if "短袖" in name or "T恤" in name:
        pt = "T-Shirt"
    elif "短裤" in name:
        pt = "Shorts"
    elif "外套" in name or "夹克" in name:
        pt = "Jacket"
    elif "Polo" in name:
        pt = "Polo Shirt"
    elif "长裤" in name or "运动套装长裤" in name:
        pt = "Track Pants"
    elif "衬衫" in name:
        pt = "Shirt"
    else:
        pt = "T-Shirt"

    return f"{brand} {pt} #{index:02d}" if brand else f"{name} #{index:02d}"


def create_excel(data: list, output_path: Path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return False

    tmp_dir = f".dumate/xlsx-{uuid.uuid4()}"
    os.makedirs(tmp_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    hf = Font(bold=True, size=11, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="4472C4")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "No.",
        "Product Name",
        "English Product Name",
        "Album ID",
        "Image Count",
        "First Image",
        "Second Image",
        "Image Links (Max 12)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.border = border

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["H"].width = 80

    for i, p in enumerate(data, 1):
        imgs = p.get("images", [])
        ws.cell(row=i + 1, column=1, value=i).border = border
        ws.cell(row=i + 1, column=2, value=p["name"]).border = border
        ws.cell(row=i + 1, column=3, value=p.get("english_name", "")).border = border
        ws.cell(row=i + 1, column=4, value=p.get("album_id", "")).border = border
        ws.cell(row=i + 1, column=5, value=len(imgs)).border = border
        ws.cell(row=i + 1, column=6, value=imgs[0] if imgs else "").border = border
        ws.cell(
            row=i + 1, column=7, value=imgs[1] if len(imgs) > 1 else ""
        ).border = border
        c = ws.cell(row=i + 1, column=8, value="\n".join(imgs[2:14]))
        c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="top")

    tmp_file = os.path.join(tmp_dir, "products.xlsx")
    wb.save(tmp_file)
    shutil.copy2(tmp_file, str(output_path))
    shutil.rmtree(tmp_dir, ignore_errors=True)
    return True


def main():
    print("=" * 60)
    print("Yupoo Product Extractor")
    print("=" * 60)

    date_str = datetime.now().strftime("%Y-%m-%d")
    json_path = OUTPUT_DIR / f"yesterday_products_{date_str}.json"
    excel_path = OUTPUT_DIR / f"yesterday_products_{date_str}.xlsx"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not Path(AUTH_FILE).exists():
        print(f"Error: {AUTH_FILE} not found")
        return 1

    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(storage_state=str(AUTH_FILE))
        page = context.new_page()

        print("Navigating to gallery...")
        page.goto("https://x.yupoo.com/gallery")
        time.sleep(3)

        print("Sorting by creation time...")
        try:
            page.click(".Dropdowns__main > a")
            time.sleep(0.5)
            page.click("text=创建时间")
            time.sleep(2)
        except:
            print("Warning: Could not sort")

        # Get product names from page text
        print("Extracting product names...")
        page_text = page.evaluate("document.body.innerText")

        # Parse product names
        product_names = []
        for line in page_text.split("\n"):
            # Match patterns like "h250 Balenciaga..." or "17 h250..."
            if re.match(r"^[hH]?\d+\s+[A-Z]", line) and len(line) > 15:
                # Clean up the name
                name = re.sub(r"^\d+\s+", "", line)  # Remove leading number count
                if name and len(name) > 10:
                    product_names.append(name)

        # Remove duplicates while preserving order
        seen = set()
        product_names = [x for x in product_names if not (x in seen or seen.add(x))]
        product_names = product_names[:30]

        print(f"Found {len(product_names)} products")

        for i, name in enumerate(product_names):
            print(f"\n[{i + 1}/{len(product_names)}] {name[:40]}...")

            try:
                # Click on product by text
                page.click(f'text="{name[:30]}"')
                time.sleep(3)

                # Get album ID from URL
                url = page.url
                album_id = ""
                m = re.search(r"gallery/(\d+)", url)
                if m:
                    album_id = m.group(1)

                # Get accurate name from title
                title = page.title()
                m = re.search(r"相册 > (.+?) \|", title)
                actual_name = m.group(1) if m else name

                # Extract images
                images = page.evaluate("""() => {
                    const imgs = [];
                    document.querySelectorAll('img').forEach(i => {
                        if (i.src.includes('photo.yupoo')) {
                            const m = i.src.match(/photo\\.yupoo\\.com\\/([^\\/]+)\\/([^\\/]+)\\//);
                            if (m) imgs.push('http://pic.yupoo.com/' + m[1] + '/' + m[2] + '/big.jpeg');
                        }
                    });
                    return [...new Set(imgs)];
                }""")

                results.append(
                    {
                        "name": actual_name,
                        "album_id": album_id,
                        "images": images or [],
                        "image_count": len(images) if images else 0,
                        "english_name": generate_english_name(actual_name, i + 1),
                    }
                )

                print(f"  Album: {album_id}, Images: {len(images) if images else 0}")

                # Go back to gallery
                page.go_back()
                time.sleep(2)

                # Save progress
                if (i + 1) % 10 == 0:
                    with open(json_path, "w", encoding="utf-8") as f:
                        json.dump(results, f, ensure_ascii=False, indent=2)

            except Exception as e:
                print(f"  Error: {e}")
                # Try to recover by going back to gallery
                try:
                    page.goto("https://x.yupoo.com/gallery")
                    time.sleep(2)
                except:
                    pass

        browser.close()

    # Save results
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    if results and create_excel(results, excel_path):
        print("\n" + "=" * 60)
        print("Extraction Complete!")
        print("=" * 60)
        print(f"JSON: {json_path}")
        print(f"Excel: {excel_path}")
        print(f"Products: {len(results)}")
        print(f"Images: {sum(r['image_count'] for r in results)}")
    else:
        print("\nNo products extracted")

    return 0


if __name__ == "__main__":
    sys.exit(main())
