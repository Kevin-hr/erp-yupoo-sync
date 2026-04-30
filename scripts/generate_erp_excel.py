#!/usr/bin/env python3
"""
Generate ERP Excel from existing products data.
Follows ERP_EXCEL_STANDARD_BAPE0418.md v2.0.0
"""

import json
import re
from pathlib import Path
from datetime import datetime
import uuid
import shutil

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    import subprocess

    subprocess.run(["pip", "install", "openpyxl"], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

# Constants from standard
CONSTANTS = {
    "H": "材质|棉质",
    "I": "N",
    "J": "Clothing",
    "M": "件/个",
    "O": "Y",
    "P": 0.3,
    "AD": 59,
    "AE": 99,
    "AF": 999,
    "Y": "Size\nS\nM\nL\nXL",
}

# Brand URL mapping
BRAND_SLUGS = {
    "Louis Vuitton": "Louis-Vuitton",
    "Balenciaga": "Balenciaga",
    "CLOT": "CLOT",
    "Saint Laurent": "Saint-Laurent",
    "Celine": "Celine",
    "Prada": "Prada",
    "Fendi": "Fendi",
    "Dior": "Dior",
    "Loewe": "Loewe",
    "Dolce & Gabbana": "Dolce-Gabbana",
    "Givenchy": "Givenchy",
    "Burberry": "Burberry",
    "Gucci": "Gucci",
    "Chanel": "Chanel",
    "Thom Browne": "Thom-Browne",
    "Moncler": "Moncler",
    "MIUMIU": "MIUMIU",
    "Maison Margiela": "Maison-Margiela",
    "OFF-WHITE": "OFF-WHITE",
    "OW": "OFF-WHITE",
    "WE11DONE": "WE11DONE",
    "DESCENTE": "DESCENTE",
    "KOLON": "KOLON",
    "Ralph Lauren": "Ralph-Lauren",
    "FENDI": "Fendi",
    "BAPE": "BAPE",
}


def extract_brand(title: str) -> str:
    """Extract brand name from title."""
    # Common patterns
    patterns = [
        r"^(h?\d+)?\s*([A-Z][A-Za-z\s&\-\.]+?)(?:/|\\|\s)",
        r"^([A-Z][A-Za-z\s&\-\.]+?)(?:/|\\|\s)",
    ]

    for pattern in patterns:
        match = re.match(pattern, title)
        if match:
            brand = match.group(2) if len(match.groups()) > 1 else match.group(1)
            brand = brand.strip()
            # Clean up
            if brand in BRAND_SLUGS or brand.upper() in BRAND_SLUGS:
                return brand
            # Check partial matches
            for b in BRAND_SLUGS.keys():
                if b.lower() in brand.lower() or brand.lower() in b.lower():
                    return b
            return brand

    return "Unknown"


def get_brand_slug(brand: str) -> str:
    """Get URL slug for brand."""
    # Direct match
    if brand in BRAND_SLUGS:
        return BRAND_SLUGS[brand]
    # Case insensitive match
    for b, slug in BRAND_SLUGS.items():
        if b.lower() == brand.lower():
            return slug
    # Default: replace spaces with hyphens
    return brand.replace(" ", "-").replace("/", "-")


def strip_brand(title: str, brand: str) -> str:
    """Remove brand prefix from title for Name field."""
    # Remove hxxx prefix
    title = re.sub(r"^h?\d+\s*", "", title)
    # Remove brand prefix
    for b in [brand, brand.upper(), brand.lower()]:
        patterns = [
            rf"^{re.escape(b)}[/\\]\s*",
            rf"^{re.escape(b)}\s+",
            rf"^{re.escape(b)}$",
        ]
        for pattern in patterns:
            title = re.sub(pattern, "", title, flags=re.IGNORECASE)
    return title.strip()


def build_description_html(title: str, brand: str) -> str:
    """Build D column HTML description."""
    brand_slug = get_brand_slug(brand)
    name_without_brand = strip_brand(title, brand)

    html = f"""<p><span style="font-family: Tahoma;"><span>Name: <span style="font-family: verdana, geneva, sans-serif;"><a href="https://www.stockxshoesvip.net/{brand_slug}/" rel="noopener" target="_blank">{brand}</a> {name_without_brand}</span></span></span></p>
<p>Category: <span style="font-size: 14px;"><a href="https://www.stockxshoesvip.net/{brand_slug}-Clothes/" target="_self" class="third-link animation-underline">{brand}</a></span><a href="https://www.stockxshoesvip.net/{brand_slug}-Clothes/" target="_self"> Clothes</a></p>
<p><span style="font-weight: bold;">More about&nbsp;</span><span style="font-weight: bold;"></span></p>
<p><a href="https://www.stockxshoesvip.net/Fear-Of-God-Clothing/" rel="noopener" target="_blank">Fear Of God Clothing</a>&nbsp; <a href="https://www.stockxshoesvip.net/Pants/" rel="noopener" target="_blank">Pants</a>&nbsp; <a href="https://www.stockxshoesvip.net/Denim-Tears/" rel="noopener" target="_blank">Denim Tears</a></p>
<p><b>Our Core Guarantees</b></p>
<ul>
<li><b>Exclusive <a href="https://www.stockxshoesvip.net/Stockxshoes-QC-Pics/" rel="noopener" target="_blank">QC Service</a>:</b> We provide free Quality Control (QC) pictures before shipment. You approve the exact item you will receive&mdash;if not satisfied, we offer free exchanges or refunds.</li>
<li><b>Premium Packaging:</b> All apparel comes with full brand packaging and original tags.</li>
<li><b>Worry-Free Logistics:</b> We handle secure delivery and customs clearance to ensure your package arrives safely.</li>
<li><b>100% Safe Shopping:</b> 30-day money-back guarantee with damage protection.</li>
</ul>
<p><b>Shipping &amp; Payment</b></p>
<ul>
<li><b>Delivery Time:</b> 7-18 Days (Minor 1-3 day delays are normal). Tracking number provided.</li>
<li><b>Shipping Methods:</b> FedEx / USPS / DHL / UPS / EMS / Royal Mail.</li>
<li><b>Payment Methods:</b> Credit/Debit Card, PayPal, Bank Transfer, Cash App, Zelle.</li>
</ul>
<p><b>About StockxShoesVIP</b></p>
<p>With 10 years of offline retail and 5 years of online excellence, we are your trusted source for <b>premium replica sneakers and streetwear</b>.</p>
<p><i>(Note: We are an independent supplier and not affiliated with the StockX platform. Please bookmark our official site: stockxshoesvip.net)</i></p>
<p><b>Contact Us</b></p>
<ul>
<li><b>WhatsApp/WeChat:</b> +86 189 5920 5893</li>
<li><b>Instagram:</b> @stockxshoesvip_com</li>
</ul>
<p><i>Buy with confidence, wear with confidence.</i></p>"""

    return html


def generate_erp_excel(products: list, output_path: str):
    """Generate ERP Excel file from products list."""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息 (2)"

    # Create second sheet for 计量单位
    ws2 = wb.create_sheet("计量单位")

    # Headers (row 2)
    headers = [
        "商品ID",
        "商品标题*",
        "副标题",
        "商品描述",
        "商品首图*",
        "商品其他图片",
        "关键信息",
        "属性",
        "商品上架*",
        "物流模板*",
        "类别名称",
        "标签",
        "计量单位",
        "商品备注",
        "不记库存*",
        "商品重量*",
        "包装长度",
        "包装宽度",
        "包装高度",
        "SEO标题",
        "SEO描述",
        "SEO关键词",
        "SEO URL Handle",
        "规格1",
        "规格2",
        "规格3",
        "规格4",
        "SKU值",
        "SKU图片",
        "售价*",
        "原价",
        "库存",
        "SKU",
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)

    # Row 3: descriptions (empty for most)
    descriptions = [
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    for col, desc in enumerate(descriptions, 1):
        ws.cell(row=3, column=col, value=desc)

    # Data rows start at row 4
    current_row = 4
    seen_titles = {}

    for product in products:
        title = product.get("name", product.get("title", ""))
        images = product.get("images", [])

        if not title or not images:
            continue

        # Deduplicate title
        base_title = title
        if base_title in seen_titles:
            seen_titles[base_title] += 1
            unique_title = f"{title} ({seen_titles[base_title]})"
        else:
            seen_titles[base_title] = 1
            unique_title = title

        # Limit title length
        if len(unique_title) > 255:
            unique_title = unique_title[:252] + "..."

        # Extract brand
        brand = extract_brand(title)

        # Limit images to 14
        images = images[:14]
        first_image = images[0] if images else ""
        other_images = "\n".join(images[1:]) if len(images) > 1 else ""

        # Build description HTML
        description = build_description_html(title, brand)

        # SEO fields
        seo_title = f"Stockx Replica Streetwear | Top Quality 1:1 {unique_title} - stockxshoesvip.net"
        seo_desc = f"Buy Best 1:1 Replica Clothing on Stockxshoesvip.net. Perfect {unique_title}. 100% safe shipping, free QC confirmation, and easy returns."

        # Main row (Size:S)
        row_data = [
            "",  # A: 商品ID
            unique_title,  # B: 商品标题
            "",  # C: 副标题
            description,  # D: 商品描述
            first_image,  # E: 商品首图
            other_images,  # F: 商品其他图片
            "",  # G: 关键信息
            CONSTANTS["H"],  # H: 属性
            CONSTANTS["I"],  # I: 商品上架
            CONSTANTS["J"],  # J: 物流模板
            brand,  # K: 类别名称
            unique_title,  # L: 标签
            CONSTANTS["M"],  # M: 计量单位
            "",  # N: 商品备注
            CONSTANTS["O"],  # O: 不记库存
            CONSTANTS["P"],  # P: 商品重量
            "",  # Q: 包装长度
            "",  # R: 包装宽度
            "",  # S: 包装高度
            seo_title,  # T: SEO标题
            seo_desc,  # U: SEO描述
            unique_title,  # V: SEO关键词
            "",  # W: SEO URL Handle
            "",  # X: 规格1
            CONSTANTS["Y"],  # Y: 规格2
            "",  # Z: 规格3
            "",  # AA: 规格4
            "Size:S",  # AB: SKU值
            "",  # AC: SKU图片
            CONSTANTS["AD"],  # AD: 售价
            CONSTANTS["AE"],  # AE: 原价
            CONSTANTS["AF"],  # AF: 库存
            "",  # AG: SKU
        ]

        for col, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col, value=value)

        current_row += 1

        # SKU sub-rows (Size:M, L, XL)
        for size in ["M", "L", "XL"]:
            sku_row = [""] * 33
            sku_row[27] = f"Size:{size}"  # AB: SKU值
            sku_row[29] = CONSTANTS["AD"]  # AD: 售价
            sku_row[30] = CONSTANTS["AE"]  # AE: 原价
            sku_row[31] = CONSTANTS["AF"]  # AF: 库存

            for col, value in enumerate(sku_row, 1):
                ws.cell(row=current_row, column=col, value=value)

            current_row += 1

    # Save
    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"Total rows: {current_row - 4} (products x 4)")


def main():
    # Load products
    input_file = Path("inputs/yesterday_products_final.json")
    if not input_file.exists():
        print(f"Error: {input_file} not found")
        return

    with open(input_file, "r", encoding="utf-8") as f:
        products = json.load(f)

    print(f"Loaded {len(products)} products")

    # Generate output filename
    today = datetime.now().strftime("%Y-%m-%d")
    output_file = f"inputs/ERP_Export_{today}_{len(products)}products.xlsx"

    # Generate Excel
    generate_erp_excel(products, output_file)

    print(f"\nDone! Output: {output_file}")


if __name__ == "__main__":
    main()
