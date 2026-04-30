# -*- coding: utf-8 -*-
"""
Fix duplicate product titles by adding color suffix.
Reads product images to identify color and updates titles.
"""

import json
from pathlib import Path
from collections import defaultdict

# Color mapping from image analysis
# Format: album_id -> color (extracted from image)
COLOR_MAPPING = {
    # Burberry Embroidered Chest Plaid Patch Letter T-Shirt
    "234116658": "White",  # 白色
    "234116627": "Black",  # 黑色
    # Loewe 26SS Colorful Logo Embroidered T-Shirt
    "234116587": "Black",  # 黑色
    "234116530": "White",  # 白色
    # MIU MIU Classic Embroidered Badge Letter T-Shirt
    "234116488": "Black",  # 黑色
    "234116464": "White",  # 白色
    # GUCCI Interlocking Small Logo Embroidered T-Shirt
    "234116423": "Black",  # 黑色
    "234116406": "White",  # 白色
    # GUCCI Minimal Letter Embroidered Logo T-Shirt (4 colors)
    "234116413": "White",  # 白色
    "234116325": "Black",  # 黑色
    "234116347": "White",  # 白色 (duplicate - need (2))
    "234116368": "Burgundy",  # 酒红色
    # Balenciaga Blue Circle Print Letter T-Shirt
    "234116272": "Black",  # 黑色
    "234116218": "White",  # 白色
    # Balenciaga Multi-Label Print T-Shirt
    "234116063": "Black",  # 黑色
    "234116048": "White",  # 白色
    # Balenciaga Front and Back Letter Print T-Shirt
    "234116034": "Black",  # 黑色
    "234116013": "White",  # 白色
    # Dior Patch Embroidered T-Shirt
    "234115995": "Black",  # 黑色
    "234115934": "White",  # 白色
    # Balenciaga Double B Graffiti Splash Letter T-Shirt
    "234115989": "White",  # 白色
    "234115967": "Black",  # 黑色
    # Burberry 26SS Knight Equestrian Stamp Round Neck T-Shirt
    "234115884": "White",  # 白色
    "234115841": "Black",  # 黑色
    # Off-White Blurred Oil Painting Arrow Cotton T-Shirt
    "234115466": "Black",  # 黑色
    "234115523": "White",  # 白色
    # Dior Curved Letter Logo Print T-Shirt
    "234115433": "White",  # 白色
    "234115459": "Black",  # 黑色
    # Loewe 26SS Shadow Logo Woven Embroidered T-Shirt
    "234115319": "Black",  # 黑色
    "234115321": "White",  # 白色
    # Balenciaga 26SS 3M Embroidered Manchester United Badge T-Shirt
    "234114951": "Black",  # 黑色
    "234114958": "Navy",  # 深蓝色
    # Balenciaga 26SS Full BB Monogram Jacquard T-Shirt (3 colors)
    "234114964": "Grey",  # 灰色
    "234114966": "Grey",  # 灰色 (duplicate - need (2))
    "234114973": "Black",  # 黑色
    # Burberry 26SS Small Equestrian Knight T-Shirt
    "234114217": "White",  # 白色
    "234114832": "Black",  # 黑色
    # Thom Browne Classic Four-Bar Knit Shorts
    "234114410": "Black",  # 黑色
    "234114416": "Navy",  # 深蓝色
    # Off-White Airport Tape Arrow Print T-Shirt
    "234114176": "Black",  # 黑色
    "234114185": "White",  # 白色
    # Louis Vuitton 26SS Dashed Triple Flower T-Shirt
    "234114129": "Black",  # 黑色
    "234114162": "White",  # 白色
    # Burberry Large Plaid Knit Shorts
    "234113896": "Green",  # 绿色
    "234113906": "Beige",  # 米色
    # Burberry Plaid Knit Shorts
    "234113879": "Beige",  # 米色
    "234113887": "Navy",  # 深蓝色
    # Burberry Classic Plaid Shorts
    "234113833": "Beige",  # 米色
    "234113863": "Navy",  # 深蓝色
    # Louis Vuitton 26SS Bone Logo T-Shirt
    "234113663": "Black",  # 黑色
    "234113690": "White",  # 白色
    # PRADA Headphone Letter Pattern Digital Print T-Shirt
    "234113521": "Black",  # 黑色
    "234113523": "White",  # 白色
    # Alexander Wang 26SS Red Letter T-Shirt (3 colors)
    "234113026": "White",  # 白色
    "234113102": "White",  # 白色 (duplicate - need (2))
    "234113158": "Black",  # 黑色
    # Alexander Wang 26SS Rhinestone Letter T-Shirt
    "234112975": "Pink",  # 粉色
    "234113063": "Black",  # 黑色
    # Moncler Large Chest Logo Embroidered T-Shirt
    "234112846": "White",  # 白色
    "234112852": "Black",  # 黑色
    # Louis Vuitton Graffiti Large Logo Print T-Shirt
    "234112409": "Black",  # 黑色
    "234112414": "White",  # 白色
    # Dior CD Monogram Webbing Drawstring Shorts (4 colors)
    "234115730": "Navy",  # 深蓝色
    "234115626": "Green",  # 绿色
    "234115680": "Grey",  # 灰色
    "234115777": "Light Blue",  # 浅蓝色
    # Ami Paris 26SS Color-Block Spliced T-Shirt (3 colors)
    "234111824": "White/Black",  # 白色主体黑色袖子
    "234111865": "White/Navy",  # 白色主体深蓝袖子
    "234111903": "White/Green",  # 白色主体绿色袖子
    # Burberry 26SS Functional Style Embroidered Equestrian Knight Collar Jacket
    "234111621": "Army Green",  # 军绿色
    "234111639": "Black",  # 黑色
    # Louis Vuitton 26SS Monogram Reversible Hooded Jacket
    "233873051": "Green/Black",  # 绿色/黑色
    "233873080": "Grey/Blue",  # 灰色/蓝色
}


def main():
    # Load products
    products_path = Path("inputs/yesterday_products_final.json")
    with open(products_path, "r", encoding="utf-8") as f:
        products = json.load(f)

    # Build album_id -> product mapping
    album_to_product = {}
    for p in products:
        album_id = str(p.get("album_id", ""))
        if album_id:
            album_to_product[album_id] = p

    # Load current Excel data
    from openpyxl import load_workbook

    excel_path = Path("inputs/ERP_Export_2026-04-28_96products_with_english.xlsx")
    wb = load_workbook(excel_path)

    # Get the sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "商品信息" in name:
            sheet_name = name
            break
    sheet = wb[sheet_name]

    # Build album_id -> row mapping
    album_to_row = {}
    row_idx = 4
    for p in products:
        album_id = str(p.get("album_id", ""))
        if album_id:
            album_to_row[album_id] = row_idx
        row_idx += 4

    # Track titles with colors to detect remaining duplicates
    title_color_map = defaultdict(list)  # title -> [(row, color, album_id)]

    # First pass: assign colors and collect titles
    for album_id, color in COLOR_MAPPING.items():
        if album_id in album_to_row:
            row = album_to_row[album_id]
            current_title = sheet[f"B{row}"].value
            if current_title:
                title_color_map[current_title].append((row, color, album_id))

    # Second pass: update titles with colors and handle duplicates
    updates = []
    for title, items in title_color_map.items():
        color_count = defaultdict(int)
        for row, color, album_id in items:
            color_count[color] += 1

        for row, color, album_id in items:
            base_title = title

            # Add color to title
            if color_count[color] > 1:
                # Multiple items with same color - need numbering
                color_count[color] += 0.5  # Hack to track which one we're on
                suffix_num = int(color_count[color])
                new_title = f"{base_title} {color} ({suffix_num})"
            else:
                new_title = f"{base_title} {color}"

            updates.append(
                {
                    "row": row,
                    "album_id": album_id,
                    "old_title": title,
                    "new_title": new_title,
                    "color": color,
                }
            )

    # Apply updates
    print(f"Updating {len(updates)} product titles with colors...")
    for update in updates:
        row = update["row"]
        new_title = update["new_title"]
        sheet[f"B{row}"] = new_title
        # Also update L (标签) and V (SEO关键词) which should match B
        sheet[f"L{row}"] = new_title
        sheet[f"V{row}"] = new_title

    # Save
    output_path = Path("inputs/ERP_Export_2026-04-28_96products_final.xlsx")
    wb.save(output_path)
    print(f"Saved to: {output_path}")

    # Print summary
    print("\n=== Title Updates Summary ===")
    for update in sorted(updates, key=lambda x: x["row"])[:20]:
        print(f"Row {update['row']}: {update['old_title'][:40]}...")
        print(f"         -> {update['new_title']}")
        print()

    # Verify no duplicates remain
    print("\n=== Checking for remaining duplicates ===")
    title_counts = defaultdict(int)
    for row in range(4, sheet.max_row + 1, 4):
        title = sheet[f"B{row}"].value
        if title:
            title_counts[title] += 1

    duplicates = {t: c for t, c in title_counts.items() if c > 1}
    if duplicates:
        print(f"WARNING: {len(duplicates)} duplicate titles remain!")
        for title, count in duplicates.items():
            print(f"  [{count}x] {title}")
    else:
        print("✓ All titles are unique!")


if __name__ == "__main__":
    main()
