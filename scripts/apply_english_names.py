# -*- coding: utf-8 -*-
"""
Apply English names to ERP Excel column C (副标题).
Uses openpyxl to update specific cells.
"""

import json
import shutil
from pathlib import Path
from openpyxl import load_workbook


def main():
    # Paths
    source_excel = Path("inputs/ERP_Export_2026-04-28_96products.xlsx")
    updates_json = Path("inputs/english_names_updates.json")
    output_excel = Path("inputs/ERP_Export_2026-04-28_96products_with_english.xlsx")

    # Load updates
    with open(updates_json, "r", encoding="utf-8") as f:
        updates = json.load(f)

    print(f"Loading Excel file: {source_excel}")
    wb = load_workbook(source_excel)

    # Get the sheet (note: sheet name may have suffix due to copy)
    sheet_name = None
    for name in wb.sheetnames:
        if "商品信息" in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
    print(f"Using sheet: {sheet_name}")
    sheet = wb[sheet_name]

    # Apply updates
    updated_count = 0
    for update in updates:
        cell = update["cell"]
        value = update["value"]
        sheet[cell] = value
        updated_count += 1

    # Save to new file
    wb.save(output_excel)
    print(f"Updated {updated_count} cells in column C")
    print(f"Saved to: {output_excel}")

    # Verify a few cells
    print("\n=== Verification ===")
    test_cells = ["C4", "C60", "C72", "C348"]
    for cell in test_cells:
        value = sheet[cell].value
        print(
            f"  {cell}: {value[:50]}..."
            if value and len(str(value)) > 50
            else f"  {cell}: {value}"
        )


if __name__ == "__main__":
    main()
