#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP 搜索抓取 v3 - 正确填写商品名称搜索框 + 遍历翻页查找目标
"""

import asyncio
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import async_playwright

ROOT_DIR = Path(__file__).parent.parent
COOKIES_FILE = ROOT_DIR / "logs" / "cookies.json"
TEMPLATE_FILE = ROOT_DIR / "商品导入模板.xlsx"
OUTPUT_DIR = ROOT_DIR / "output"
SCREENSHOT_DIR = ROOT_DIR / "screenshots"
LOG_DIR = ROOT_DIR / "logs"
for d in [OUTPUT_DIR, LOG_DIR, SCREENSHOT_DIR]: d.mkdir(exist_ok=True)

logger = logging.getLogger("erp_v3")
logger.setLevel(logging.INFO)
fh = logging.FileHandler(LOG_DIR / f"erp_v3_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log", encoding='utf-8')
logger.addHandler(fh)
sh = logging.StreamHandler(); logger.addHandler(sh)

ERP_BASE = "https://www.mrshopplus.com"
PRODUCT_LIST_URL = f"{ERP_BASE}/#/product/list_DTB_proProduct"


class ERPProduct:
    def __init__(self):
        self.product_id = ""; self.title = ""; self.subtitle = ""
        self.description = ""; self.primary_image = ""; self.other_images = ""
        self.key_info = ""; self.attributes = ""; self.listed = "N"
        self.shipping_template = "默认模板"; self.category = ""
        self.tags = ""; self.unit = "件/个"; self.notes = ""
        self.no_inventory = "Y"; self.weight = 0.5
        self.pack_length = 0.0; self.pack_width = 0.0; self.pack_height = 0.0
        self.seo_title = ""; self.seo_description = ""; self.seo_keywords = ""
        self.seo_handle = ""; self.spec1_name = ""; self.spec2_name = ""
        self.spec3_name = ""; self.spec4_name = ""; self.sku_values = ""
        self.sku_images = ""; self.price = 99.0; self.original_price = 199.0
        self.stock = 100; self.sku_code = ""

    def to_row(self):
        return [
            self.product_id, self.title, self.subtitle, self.description,
            self.primary_image, self.other_images, self.key_info, self.attributes,
            self.listed, self.shipping_template, self.category, self.tags, self.unit,
            self.notes, self.no_inventory, self.weight, self.pack_length,
            self.pack_width, self.pack_height, self.seo_title, self.seo_description,
            self.seo_keywords, self.seo_handle, self.spec1_name, self.spec2_name,
            self.spec3_name, self.spec4_name, self.sku_values, self.sku_images,
            self.price, self.original_price, self.stock, self.sku_code
        ]


def fill_excel(product, output_path):
    shutil.copy(TEMPLATE_FILE, output_path)
    wb = load_workbook(output_path)
    ws = wb["商品信息"]
    row = 3
    for r in range(1, 5):
        val = ws.cell(row=r, column=2).value
        if val and "商品标题" in str(val):
            row = r + 1; break
    for col_idx, value in enumerate(product.to_row(), start=1):
        ws.cell(row=row, column=col_idx, value=value)
    wb.save(output_path)


def log(msg):
    logger.info(msg)
    print(f"[ERP-v3] {msg}")


async def get_all_inputs(page):
    """获取所有可见input的详细信息"""
    return await page.evaluate("""
        () => {
            const result = [];
            document.querySelectorAll('.el-input__inner').forEach((el, i) => {
                if (el.offsetParent !== null) {
                    result.push({
                        index: i,
                        placeholder: el.placeholder,
                        value: el.value,
                        type: el.type,
                        className: el.className.substring(0, 80)
                    });
                }
            });
            return result;
        }
    """)


async def get_table_rows_full(page):
    """获取表格所有行的完整数据"""
    return await page.evaluate("""
        () => {
            const table = document.querySelector('.el-table__body-wrapper');
            if (!table) return [];
            const trs = table.querySelectorAll('tr.el-table__row');
            return Array.from(trs).map((tr, i) => {
                const cells = tr.querySelectorAll('td');
                const imgs = [];
                cells.forEach(c => {
                    c.querySelectorAll('img').forEach(img => {
                        if (img.src && !img.src.includes('logo') && !img.src.includes('head-logo')) {
                            imgs.push(img.src);
                        }
                    });
                });
                return {
                    index: i,
                    // 列：0=空, 1=商品图片, 2=商品名称, 3=时间, 4=售价, 5=原价, 6=库存, 7=?, 8=上架?, 9=?
                    name: cells[2] ? cells[2].innerText.trim().substring(0, 80) : '',
                    price: cells[4] ? cells[4].innerText.trim() : '',
                    original_price: cells[5] ? cells[5].innerText.trim() : '',
                    stock: cells[6] ? cells[6].innerText.trim() : '',
                    listed: cells[8] ? cells[8].innerText.trim() : '',
                    imgs: imgs,
                    fullText: tr.innerText.trim().substring(0, 200)
                };
            });
        }
    """)


async def get_pager_info(page):
    """获取分页信息"""
    return await page.evaluate("""
        () => {
            const pager = document.querySelector('.el-pagination');
            if (!pager) return { total: 0, current: 1, pages: 0, hasNext: false };

            // 总数
            const totalEl = pager.querySelector('.el-pagination__total');
            const totalText = totalEl ? totalEl.innerText : '';
            const totalMatch = totalText.match(/\\d+/);
            const total = totalMatch ? parseInt(totalMatch[0]) : 0;

            // 当前页
            const activeBtn = pager.querySelector('.el-pager li.is-active');
            const current = activeBtn ? parseInt(activeBtn.innerText) : 1;

            // 总页数
            const pageCount = Math.ceil(total / 15);

            // 是否有下一页
            const nextBtn = pager.querySelector('.btn-next');
            const hasNext = nextBtn ? !nextBtn.classList.contains('disabled') : false;

            return { total, current, pageCount, hasNext };
        }
    """)


async def search_and_find_product(page, keyword: str):
    """
    核心搜索逻辑：
    1. 找到商品名称搜索框（placeholder含"商品名称"或"名称"）
    2. 清空并填写关键词
    3. 点击查询
    4. 检查结果
    """
    log(f"Searching for: {keyword}")

    # 方法1: 用 Playwright locator 直接定位搜索框
    try:
        # 方法1a: placeholder 精确定位
        search_box = page.locator("input[placeholder='请输入商品名称/编码']")
        if await search_box.is_visible(timeout=3000):
            await search_box.clear()
            await search_box.fill(keyword)
            log("Filled via locator placeholder")
        else:
            raise Exception("not visible")
    except:
        try:
            # 方法1b: placeholder含"商品名称"
            search_box = page.locator(".el-input__inner").filter(placeholder=re.compile("商品名称|名称"))
            if await search_box.is_visible(timeout=3000):
                await search_box.clear()
                await search_box.fill(keyword)
                log("Filled via filter placeholder")
            else:
                raise Exception("not visible")
        except:
            try:
                # 方法1c: 直接用index（col-2）
                search_box = page.locator(".el-input__inner").nth(1)  # 第2个
                if await search_box.is_visible(timeout=3000):
                    await search_box.clear()
                    await search_box.fill(keyword)
                    log("Filled via nth(1)")
                else:
                    raise Exception("not visible")
            except Exception as e:
                log(f"ALL SEARCH METHODS FAILED: {e}")
                # 方法1d: 兜底 - JS直接操作
                await page.evaluate(f"""
                    () => {{
                        const inputs = document.querySelectorAll('.el-input__inner');
                        // 找第2个（商品名称输入框）
                        if (inputs[1]) {{
                            inputs[1].value = '{keyword}';
                            inputs[1].dispatchEvent(new Event('input', {{ bubbles: true }}));
                            inputs[1].dispatchEvent(new Event('change', {{ bubbles: true }}));
                        }}
                    }}
                """)
                log("Filled via JS nth(1)")

    await asyncio.sleep(1)

    # 点击查询按钮
    try:
        await page.get_by_role("button", name="查询").click()
        log("Clicked 查询")
    except:
        await page.evaluate("""
            () => {
                const btns = document.querySelectorAll('button');
                for (const btn of btns) {
                    if (btn.innerText.trim() === '查询') {
                        btn.click();
                        break;
                    }
                }
            }
        """)
        log("Clicked 查询 via JS")

    await asyncio.sleep(4)  # 等待搜索结果

    # 检查结果
    rows = await get_table_rows_full(page)
    pager = await get_pager_info(page)
    log(f"Result: {pager['total']} total, {len(rows)} rows on current page")
    for r in rows:
        log(f"  Row {r['index']}: {r['name']} | price={r['price']} | orig={r['original_price']} | imgs={len(r['imgs'])}")

    return rows, pager


async def go_to_page(page, page_num: int):
    """翻到指定页"""
    js = (
        "() => {"
        "  var pager = document.querySelector('.el-pagination');"
        "  if (!pager) return;"
        "  var btns = pager.querySelectorAll('.el-pager li.number');"
        "  var target = '%s';"
        "  for (var i = 0; i < btns.length; i++) {"
        "    if (btns[i].innerText.trim() === target) { btns[i].click(); break; }"
        "  }"
        "}"
    ) % page_num
    await page.evaluate(js)
    await asyncio.sleep(3)


async def click_product_row(page, keyword: str):
    """点击指定商品行，进入详情"""
    await page.evaluate(f"""
        (kw) => {{
            const table = document.querySelector('.el-table__body-wrapper');
            if (!table) return;
            const trs = table.querySelectorAll('tr.el-table__row');
            for (const tr of trs) {{
                const cells = tr.querySelectorAll('td');
                const name = cells[2] ? cells[2].innerText.trim() : '';
                if (name.toLowerCase().includes(kw.toLowerCase())) {{
                    // 点击商品名称单元格
                    if (cells[2]) cells[2].click();
                    else if (cells[0]) cells[0].click();
                    break;
                }}
            }}
        }}
    """, keyword)
    await asyncio.sleep(5)


async def extract_from_detail_page(page, keyword: str) -> dict:
    """从商品详情页提取完整数据"""
    await asyncio.sleep(3)

    # 获取 URL
    url = page.url
    log(f"Detail URL: {url}")

    # 截图
    await page.screenshot(path=str(SCREENSHOT_DIR / "erp_v3_detail.png"), timeout=10000)

    # 从 DOM 提取
    data = await page.evaluate("""
        () => {
            const result = {
                url: window.location.href,
                price: null,
                originalPrice: null,
                stock: null,
                title: null,
                category: null,
                tags: null,
                imgs: [],
                desc: '',
                weight: null
            };

            // 价格
            const inputs = document.querySelectorAll('input');
            inputs.forEach(inp => {
                const ph = (inp.placeholder || '').replace(/\\s/g, '');
                const val = inp.value;
                if (ph.includes('售价') && val) result.price = val;
                if ((ph.includes('原价') || ph.includes('划线')) && val) result.originalPrice = val;
                if (ph.includes('库存') && val) result.stock = val;
                if ((ph.includes('商品名称') || ph.includes('标题')) && val) result.title = val;
                if (ph.includes('重量') && val) result.weight = val;
            });

            // 分类
            const selects = document.querySelectorAll('.el-select');
            selects.forEach(sel => {
                const label = sel.querySelector('.el-select__label');
                if (label) {
                    const text = label.innerText.trim();
                    if (text && !result.category) result.category = text;
                }
            });

            // 图片
            document.querySelectorAll('img').forEach(img => {
                const src = img.src || '';
                if (src && !src.includes('logo') && !src.includes('head-logo')
                    && !src.includes('empty') && img.offsetParent !== null) {
                    result.imgs.push(src);
                }
            });

            // TinyMCE
            const iframes = document.querySelectorAll('iframe');
            for (const frame of iframes) {
                try {
                    const body = frame.contentDocument.body;
                    if (body && body.innerText && body.innerText.trim()) {
                        result.desc = body.innerText.trim();
                        break;
                    }
                } catch(e) {}
            }

            // 商品名称（另一种方式）
            const titleInputs = document.querySelectorAll('input');
            titleInputs.forEach(inp => {
                if (inp.value && inp.value.length > 5 && inp.value.length < 200) {
                    if (!result.title) result.title = inp.value;
                }
            });

            return result;
        }
    """)

    log(f"Detail data: price={data['price']}, orig={data['originalPrice']}, title={data['title']}")
    log(f"Detail images: {data['imgs'][:3]}")
    log(f"Detail desc: {data['desc'][:80] if data['desc'] else 'empty'}")

    return data


async def main(keyword: str):
    log(f"=== ERP Search & Extract v3: {keyword} ===")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={"width": 1920, "height": 1080})

        if COOKIES_FILE.exists():
            cookies = json.load(open(COOKIES_FILE, 'r', encoding='utf-8'))
            await context.add_cookies(cookies)
            log(f"Loaded {len(cookies)} cookies")

        page = await context.new_page()

        # ===== 导航到商品列表 =====
        await page.goto(PRODUCT_LIST_URL, timeout=30000)
        await asyncio.sleep(6)

        # 截图初始状态
        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_v3_initial.png"), timeout=10000)

        # ===== 获取分页信息 =====
        pager = await get_pager_info(page)
        log(f"Pager: total={pager['total']}, current={pager['current']}, pages={pager['pageCount']}")

        # ===== 尝试搜索 =====
        rows, pager = await search_and_find_product(page, keyword)

        # 检查搜索结果
        matched = None
        for r in rows:
            if keyword.lower() in r['name'].lower():
                matched = r
                log(f"MATCHED! {r['name']}")
                break

        # 如果没搜到，翻页查找
        if not matched and pager['total'] > len(rows):
            log(f"Searching across pages (total={pager['total']})...")
            # 翻页遍历（最多看20页）
            for page_num in range(1, min(21, pager['pageCount'] + 1)):
                if page_num != pager['current']:
                    await go_to_page(page, page_num)
                rows = await get_table_rows_full(page)
                pager = await get_pager_info(page)
                for r in rows:
                    if keyword.lower() in r['name'].lower():
                        matched = r
                        log(f"MATCHED on page {page_num}! {r['name']}")
                        break
                if matched:
                    break
                else:
                    log(f"  Page {page_num}: no match")

        # 如果找到了，提取详情
        product = ERPProduct()
        product.title = keyword

        if matched:
            log(f"Building product from matched row: {matched['name']}")
            product.title = matched['name']

            # 从行数据提取
            try:
                nums = re.findall(r'[\d.]+', matched.get('price', ''))
                if nums: product.price = float(nums[0])
            except: pass
            try:
                nums = re.findall(r'[\d.]+', matched.get('original_price', ''))
                if nums: product.original_price = float(nums[0])
            except: pass
            try:
                nums = re.findall(r'\d+', matched.get('stock', ''))
                if nums: product.stock = int(nums[0])
            except: pass
            product.listed = 'Y' if '上架' in matched.get('listed', '') or matched.get('listed', '') == 'Y' else 'N'

            # 图片
            for img in matched.get('imgs', []):
                if not product.primary_image:
                    product.primary_image = img
                else:
                    product.other_images = (product.other_images + '\n' + img) if product.other_images else img

            # 点击进入详情
            await click_product_row(page, keyword)
            detail = await extract_from_detail_page(page, keyword)

            if detail:
                if detail.get('title'): product.title = detail['title']
                if detail.get('price'):
                    try: product.price = float(detail['price'])
                    except: pass
                if detail.get('originalPrice'):
                    try: product.original_price = float(detail['originalPrice'])
                    except: pass
                if detail.get('category'): product.category = detail['category']
                if detail.get('tags'): product.tags = detail['tags']
                if detail.get('weight'):
                    try: product.weight = float(detail['weight'])
                    except: pass
                if detail.get('desc') and not product.description:
                    product.description = f"<p>{detail['desc']}</p>"

                # 合并图片
                if detail.get('imgs'):
                    existing = ([product.primary_image] if product.primary_image else []) + \
                               (product.other_images.split('\n') if product.other_images else [])
                    for img in detail['imgs']:
                        if img and img not in existing:
                            existing.append(img)
                    if existing:
                        product.primary_image = existing[0]
                        product.other_images = '\n'.join(existing[1:15])

        else:
            log(f"Product NOT found after searching all pages. Using keyword as title.")
            product.title = keyword

        # SEO Handle
        product.seo_handle = re.sub(r'[^a-z0-9]+', '-', product.title.lower())
        product.seo_handle = re.sub(r'^-+|-+$', '', product.seo_handle)[:100]

        # ===== 写 Excel =====
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = re.sub(r'[\\/:*?"<>|]', '', keyword[:20])
        output_path = OUTPUT_DIR / f"ERP提取_{safe}_{ts}.xlsx"
        fill_excel(product, output_path)

        log(f"=== DONE ===")
        log(f"Title: {product.title}")
        log(f"Price: {product.price} / Original: {product.original_price}")
        log(f"Stock: {product.stock}")
        log(f"Listed: {product.listed}")
        log(f"Primary Image: {product.primary_image[:80] if product.primary_image else 'NONE'}")
        log(f"Other Images count: {len(product.other_images.split(chr(10))) if product.other_images else 0}")
        log(f"Description: {product.description[:100] if product.description else 'empty'}...")
        log(f"Output: {output_path}")

        await asyncio.sleep(2)
        await browser.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("keyword", help="商品关键词")
    args = parser.parse_args()
    asyncio.run(main(args.keyword))
