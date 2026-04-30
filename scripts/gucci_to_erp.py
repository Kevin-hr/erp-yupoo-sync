#!/usr/bin/env python3
"""
Gucci 图片链接表 → ERP 标准导入模板转换器

关键约束 (P0):
  - I列必须 N (下架)
  - 单商品图片总数 ≤ 14 (含首图)
  - E列单URL, F列换行分隔多URL
  - P列重量为 kg，3位小数
  - AD/AE 为 2位小数
  - J列物流模板必须是系统已配置模板

输出遵循 templates/商品导入模板 (修改版1.0).xlsx 的结构与样式，
每个商品生成 4 行 SKU (Size:S/M/L/XL)。
"""

import argparse
from copy import copy
import re
import sys
from pathlib import Path
from typing import DefaultDict, Iterable, List, Optional, Tuple
from collections import Counter, defaultdict

import openpyxl

COL_ID = 1  # A
COL_TITLE = 2  # B
COL_SUBTITLE = 3  # C
COL_DESC = 4  # D
COL_MAIN_IMG = 5  # E
COL_OTHER_IMG = 6  # F
COL_KEYINFO = 7  # G
COL_ATTR = 8  # H
COL_PUBLISH = 9  # I
COL_LOGISTICS = 10  # J
COL_CATEGORY = 11  # K
COL_TAGS = 12  # L
COL_UNIT = 13  # M
COL_NOTE = 14  # N
COL_NO_INV = 15  # O
COL_WEIGHT = 16  # P
COL_LENGTH = 17  # Q
COL_WIDTH = 18  # R
COL_HEIGHT = 19  # S
COL_SEO_TITLE = 20  # T
COL_SEO_DESC = 21  # U
COL_SEO_KEY = 22  # V
COL_SEO_URL = 23  # W
COL_SPEC1 = 24  # X
COL_SPEC2 = 25  # Y
COL_SPEC3 = 26  # Z
COL_SPEC4 = 27  # AA
COL_SKU_VAL = 28  # AB
COL_SKU_IMG = 29  # AC
COL_PRICE = 30  # AD
COL_ORIG_PRICE = 31  # AE
COL_INVENTORY = 32  # AF
COL_SKU_CODE = 33  # AG

SIZE_LIST = ["S", "M", "L", "XL"]


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub("[^a-z0-9\\s-]", "", text)
    text = re.sub("[\\s]+", "-", text)
    text = re.sub("-+", "-", text).strip("-")
    return f"gucci-{text}" if text else "gucci-item"


def _escape_html(s: object) -> str:
    s = "" if s is None else str(s)
    s = s.replace("&", "&amp;")
    s = s.replace("<", "&lt;")
    s = s.replace(">", "&gt;")
    s = s.replace('"', "&quot;")
    s = s.replace("'", "&apos;")
    return s


def _strip_brand(title: str, brand: str) -> str:
    t = str(title or "").strip()
    b = str(brand or "").strip()
    if not t or not b:
        return t
    return re.sub(rf"^{re.escape(b)}\s*[/]?\s*", "", t, flags=re.IGNORECASE).strip()


def _build_desc_html(title: str, brand: str) -> str:
    name_field = _strip_brand(title, brand)
    brand_hyphen = brand.replace(" ", "-")
    html = (
        f'<p><span style="font-family: Tahoma;"><span>Name: '
        f'<span style="font-family: verdana, geneva, sans-serif;">'
        f'<a href="https://www.stockxshoesvip.net/{brand_hyphen}/" '
        f'rel="noopener" target="_blank">{_escape_html(brand)}</a> {_escape_html(name_field)}'
        f'</span></span></span></p>'
        f'<p>Category: <span style="font-size: 14px;">'
        f'<a href="https://www.stockxshoesvip.net/{brand_hyphen}-Clothes/" '
        f'target="_self" class="third-link animation-underline">{_escape_html(brand)} Clothes</a>'
        f"</span></p>"
        f"<p><b>Our Core Guarantees</b></p>"
        f"<ul>"
        f'<li><b>Exclusive <a href="https://www.stockxshoesvip.net/Stockxshoes-QC-Pics/" '
        f'rel="noopener" target="_blank">QC Service</a>:</b> We provide free Quality Control (QC) '
        f"pictures before shipment. You approve the exact item you will receive.</li>"
        f"<li><b>Premium Packaging:</b> All apparel comes with full brand packaging and original tags.</li>"
        f"<li><b>Worry-Free Logistics:</b> Secure delivery and customs clearance.</li>"
        f"<li><b>100% Safe Shopping:</b> 30-day money-back guarantee.</li>"
        f"</ul>"
        f"<p><b>Shipping &amp; Payment</b></p>"
        f"<ul>"
        f"<li><b>Delivery Time:</b> 7-18 Days. Tracking number provided.</li>"
        f"<li><b>Shipping Methods:</b> FedEx / USPS / DHL / UPS / EMS / Royal Mail.</li>"
        f"<li><b>Payment Methods:</b> Credit/Debit Card, PayPal, Bank Transfer, Cash App, Zelle.</li>"
        f"</ul>"
        f"<p><b>About StockxShoesVIP</b></p>"
        f"<p>With 10 years of offline retail and 5 years of online excellence.</p>"
        f"<p><i>(Note: We are not affiliated with the StockX platform. "
        f"Please bookmark: stockxshoesvip.net)</i></p>"
        f"<p><b>Contact Us</b></p>"
        f"<ul>"
        f"<li><b>WhatsApp/WeChat:</b> +86 189 5920 5893</li>"
        f"<li><b>Instagram:</b> @stockxshoesvip_com</li>"
        f"</ul>"
        f"<p><i>Buy with confidence, wear with confidence.</i></p>"
    )
    return html


def _build_seo_title(title: str) -> str:
    return f"Stockx Replica Streetwear | Top Quality 1:1 {title} - stockxshoesvip.net"


def _build_seo_desc(title: str) -> str:
    return (
        "Buy Best 1:1 Replica Clothing on Stockxshoesvip.net. "
        f"Perfect {title}. 100% safe shipping, free QC confirmation, and easy returns."
    )


def _resource_root() -> Path:
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            return Path(meipass)
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _clear_row_values(ws, row_idx: int, max_col: int = 33) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row_idx, c).value = None


def _make_unique_title(title: str, used_titles: set, seq: Optional[int] = None) -> str:
    base = str(title or "").strip()
    if not base:
        return base

    max_len = 255
    if seq is None:
        candidate = base[:max_len]
    else:
        suffix = f" {seq:02d}"
        keep_len = max_len - len(suffix)
        candidate = f"{base[:keep_len].rstrip()}{suffix}"

    ckey = candidate.lower()
    if ckey not in used_titles:
        used_titles.add(ckey)
        return candidate

    idx = 2
    while True:
        if seq is None:
            suffix = f" ({idx})"
        else:
            suffix = f" {seq:02d}-{idx}"
        keep_len = max_len - len(suffix)
        candidate2 = f"{base[:keep_len].rstrip()}{suffix}"
        ckey2 = candidate2.lower()
        if ckey2 not in used_titles:
            used_titles.add(ckey2)
            return candidate2
        idx += 1


def _is_url(val: object) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return s.startswith("http://") or s.startswith("https://")


def _uniq_preserve_order(items: Iterable[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for it in items:
        if it in seen:
            continue
        seen.add(it)
        out.append(it)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="输入 xlsx（图片链接表）")
    ap.add_argument("--output", required=True, help="输出 ERP 模板 xlsx")
    ap.add_argument("--template", default=None, help="ERP 模板路径（默认 templates/商品导入模板 (修改版1.0).xlsx）")
    ap.add_argument("--brand", default="GUCCI", help="品牌名（默认 GUCCI）")
    args = ap.parse_args()

    root = _resource_root().parent
    template = Path(args.template) if args.template else (root / "templates" / "商品导入模板 (修改版1.0).xlsx")

    wb_in = openpyxl.load_workbook(args.input)
    ws_in = wb_in.active

    wb = openpyxl.load_workbook(template)
    ws = wb.active

    used_titles: set[str] = set()
    out_row = 4
    for r in range(2, ws_in.max_row + 1):
        title = str(ws_in.cell(r, 2).value or "").strip()
        if not title:
            continue

        main_img = str(ws_in.cell(r, 5).value or "").strip()
        others: List[str] = []
        raw_others = str(ws_in.cell(r, 6).value or "").strip()
        if raw_others:
            others = [x.strip() for x in raw_others.splitlines() if x.strip()]
        imgs = [main_img] + others
        imgs = [u.replace("photo.yupoo.com", "pic.yupoo.com") for u in imgs if _is_url(u)]
        imgs = _uniq_preserve_order(imgs)[:14]

        if not imgs:
            continue

        title2 = _make_unique_title(title, used_titles)
        main_img2 = imgs[0]
        other_img2 = "\n".join(imgs[1:])

        _clear_row_values(ws, out_row)
        ws.cell(out_row, COL_TITLE).value = title2
        ws.cell(out_row, COL_DESC).value = _build_desc_html(title2, args.brand)
        ws.cell(out_row, COL_MAIN_IMG).value = main_img2
        ws.cell(out_row, COL_OTHER_IMG).value = other_img2
        ws.cell(out_row, COL_PUBLISH).value = "N"
        ws.cell(out_row, COL_CATEGORY).value = args.brand
        ws.cell(out_row, COL_TAGS).value = title2
        ws.cell(out_row, COL_SEO_TITLE).value = _build_seo_title(title2)
        ws.cell(out_row, COL_SEO_DESC).value = _build_seo_desc(title2)
        ws.cell(out_row, COL_SEO_KEY).value = title2

        base_style_row = out_row
        out_row += 1
        for size in SIZE_LIST:
            _clear_row_values(ws, out_row)
            for col in range(1, COL_SKU_CODE + 1):
                ws.cell(out_row, col)._style = copy(ws.cell(base_style_row, col)._style)
            ws.cell(out_row, COL_SKU_VAL).value = f"Size:{size}"
            out_row += 1

    wb.save(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

