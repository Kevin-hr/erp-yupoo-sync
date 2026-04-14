#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP JavaScript 评估抓取 - 通过页面 JS 上下文提取商品数据
JS Evaluation Extractor: Use Playwright page.evaluate() to extract data from Vue SPA

关键思路：
- Vue SPA 的数据在 JavaScript 内存中，不在原始 HTML 里
- 通过 page.evaluate() 执行 JS 直接从 Vue 实例 / DOM 状态 获取数据
"""

import asyncio
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import async_playwright

# =============================================================================
# 路径
# =============================================================================
ROOT_DIR = Path(__file__).parent.parent
COOKIES_FILE = ROOT_DIR / "logs" / "cookies.json"
TEMPLATE_FILE = ROOT_DIR / "商品导入模板.xlsx"
OUTPUT_DIR = ROOT_DIR / "output"
SCREENSHOT_DIR = ROOT_DIR / "screenshots"
LOG_DIR = ROOT_DIR / "logs"

for d in [OUTPUT_DIR, LOG_DIR, SCREENSHOT_DIR]:
    d.mkdir(exist_ok=True)

logger = logging.getLogger("erp_js")
logger.setLevel(logging.INFO)
fh = logging.FileHandler(LOG_DIR / f"erp_js_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log", encoding='utf-8')
logger.addHandler(fh)
sh = logging.StreamHandler()
logger.addHandler(sh)

ERP_BASE = "https://www.mrshopplus.com"
PRODUCT_LIST_URL = f"{ERP_BASE}/#/product/list_DTB_proProduct"


class ERPProduct:
    def __init__(self):
        self.product_id = ""; self.title = ""; self.subtitle = ""
        self.description = ""; self.primary_image = ""; self.other_images = ""
        self.key_info = ""; self.attributes = ""; self.listed = "N"
        self.shipping_template = "默认模板"; self.category = ""
        self.tags = ""; self.unit = "件/个"; self.notes = ""
        self.no_inventory = "Y"; self.weight = 0.5
        self.pack_length = 0.0; self.pack_width = 0.0; self.pack_height = 0.0
        self.seo_title = ""; self.seo_description = ""; self.seo_keywords = ""
        self.seo_handle = ""; self.spec1_name = ""; self.spec2_name = ""
        self.spec3_name = ""; self.spec4_name = ""; self.sku_values = ""
        self.sku_images = ""; self.price = 99.0; self.original_price = 199.0
        self.stock = 100; self.sku_code = ""

    def to_row(self):
        return [
            self.product_id, self.title, self.subtitle, self.description,
            self.primary_image, self.other_images, self.key_info, self.attributes,
            self.listed, self.shipping_template, self.category, self.tags, self.unit,
            self.notes, self.no_inventory, self.weight, self.pack_length,
            self.pack_width, self.pack_height, self.seo_title, self.seo_description,
            self.seo_keywords, self.seo_handle, self.spec1_name, self.spec2_name,
            self.spec3_name, self.spec4_name, self.sku_values, self.sku_images,
            self.price, self.original_price, self.stock, self.sku_code
        ]


async def explore_page_structure(page) -> dict:
    """探索页面结构，打印所有输入框和可见文本"""
    info = await page.evaluate("""
        () => {
            // 找所有 input
            const inputs = [];
            document.querySelectorAll('input').forEach((el, i) => {
                inputs.push({
                    index: i,
                    tag: el.tagName,
                    type: el.type,
                    placeholder: el.placeholder,
                    value: el.value,
                    visible: el.offsetParent !== null,
                    className: el.className.substring(0, 80)
                });
            });

            // 找 el-input wrapper
            const elInputs = [];
            document.querySelectorAll('.el-input').forEach((el, i) => {
                const input = el.querySelector('input');
                if (input) {
                    elInputs.push({
                        placeholder: input.placeholder,
                        value: input.value,
                        visible: el.offsetParent !== null
                    });
                }
            });

            // 找搜索相关按钮
            const buttons = [];
            document.querySelectorAll('button').forEach(btn => {
                const txt = btn.innerText.trim();
                if (txt) buttons.push(txt.substring(0, 30));
            });

            // 找表格行
            const tableRows = document.querySelectorAll('.el-table__body-wrapper tr');
            const rowTexts = [];
            tableRows.forEach((row, i) => {
                if (i < 5) {
                    const cells = row.querySelectorAll('td');
                    const cellTexts = [];
                    cells.forEach(c => cellTexts.push(c.innerText.trim().substring(0, 30)));
                    rowTexts.push({row: i, cells: cellTexts});
                }
            });

            // Vue 组件数据（如果有）
            let vueData = null;
            const vueApp = document.querySelector('#app').__vue_app__;
            if (vueApp) {
                try {
                    const instance = document.querySelector('#app').__vue_app__._instance;
                    if (instance) {
                        vueData = 'found vue app';
                    }
                } catch(e) {
                    vueData = 'vue error: ' + e.message;
                }
            }

            return {
                inputs: inputs.filter(i => i.visible).slice(0, 10),
                elInputs: elInputs.filter(e => e.visible).slice(0, 5),
                buttons: buttons.slice(0, 10),
                rowTexts: rowTexts,
                vueData: vueData,
                url: window.location.href,
                title: document.title
            };
        }
    """)
    return info


async def search_with_js(page, keyword: str):
    """通过 JS 直接操作 Vue 组件执行搜索"""
    logger.info(f"[JS-Search] Starting JS search for: {keyword}")

    # 尝试多种方式触发搜索
    result = await page.evaluate("""
        (keyword) => {
            const results = { method: 'none', success: false, message: '' };

            // 方法1: 找 el-input 的搜索框
            const elInputs = document.querySelectorAll('.el-input__inner');
            let searchInput = null;
            elInputs.forEach(input => {
                const ph = input.placeholder || '';
                if (ph.includes('搜索') || ph.includes('商品') || ph.includes('查询')) {
                    searchInput = input;
                }
            });

            if (searchInput) {
                // 模拟用户输入
                searchInput.focus();
                searchInput.value = '';
                searchInput.dispatchEvent(new Event('input', { bubbles: true }));

                // 逐字输入
                const nativeInputValueSetter = Object.getOwnPropertyDescriptor(
                    window.HTMLInputElement.prototype, 'value'
                ).set;
                nativeInputValueSetter(searchInput, keyword);
                searchInput.dispatchEvent(new Event('input', { bubbles: true }));
                searchInput.dispatchEvent(new Event('change', { bubbles: true }));

                results.method = 'el-input-inner';
                results.success = true;
                results.message = 'Filled via el-input__inner: ' + searchInput.placeholder;

                // 尝试按回车
                document.activeElement.blur();
                return results;
            }

            // 方法2: 找任何可见 input
            const allInputs = document.querySelectorAll('input');
            allInputs.forEach(input => {
                if (input.offsetParent !== null) {
                    const ph = input.placeholder || '';
                    input.value = keyword;
                    input.dispatchEvent(new Event('input', { bubbles: true }));
                    results.method = 'any-input';
                    results.success = true;
                    results.message = 'Filled via: ' + ph + ' class: ' + input.className;
                }
            });

            // 方法3: Vue ref 直接赋值（高级）
            // 尝试找 Vue 实例并调用搜索方法
            try {
                const app = document.querySelector('#app');
                if (app && app.__vue__) {
                    const vue = app.__vue__;
                    // 尝试找 search/query 方法
                    for (const key of Object.keys(vue)) {
                        if (key.toLowerCase().includes('search') ||
                            key.toLowerCase().includes('query') ||
                            key.toLowerCase().includes('filter')) {
                            results.message += ' | Found Vue method: ' + key;
                        }
                    }
                }
            } catch(e) {
                results.message += ' | Vue access error: ' + e.message;
            }

            return results;
        }
    """, keyword)

    logger.info(f"[JS-Search] Result: {result}")
    return result


async def get_table_data_via_js(page) -> list:
    """通过 JS 从 Vue 表格组件提取数据"""
    data = await page.evaluate("""
        () => {
            // 尝试从 el-table 获取数据
            const table = document.querySelector('.el-table__body-wrapper');
            if (!table) return { error: 'No table found', rows: [] };

            const rows = [];
            const trs = table.querySelectorAll('tr.el-table__row');
            trs.forEach((tr, i) => {
                const cells = tr.querySelectorAll('td');
                const rowData = {
                    index: i,
                    cells: []
                };
                cells.forEach((cell, ci) => {
                    // 获取单元格内的文本和图片
                    const text = cell.innerText.trim();
                    const imgs = [];
                    cell.querySelectorAll('img').forEach(img => {
                        if (img.src) imgs.push(img.src);
                    });
                    rowData.cells.push({
                        text: text.substring(0, 50),
                        images: imgs.slice(0, 3)
                    });
                });
                rows.push(rowData);
            });

            // 尝试从 Vue store 获取数据
            let storeData = null;
            try {
                const elTables = document.querySelectorAll('.el-table');
                elTables.forEach(t => {
                    if (t.__vue_parent_component) {
                        const vm = t.__vue_parent_component;
                        if (vm.ctx && vm.ctx.tableData) {
                            storeData = vm.ctx.tableData.slice(0, 5);
                        } else if (vm.proxy && vm.proxy.tableData) {
                            storeData = vm.proxy.tableData.slice(0, 5);
                        }
                    }
                });
            } catch(e) {
                storeData = 'error: ' + e.message;
            }

            return { rows: rows, storeData: storeData, rowCount: len(trs) };
        }
    """)
    return data


async def click_row_and_get_detail(page, row_index: int) -> dict:
    """点击表格某行，进入编辑详情页，提取所有数据"""
    result = await page.evaluate("""
        (rowIndex) => {
            const table = document.querySelector('.el-table__body-wrapper');
            if (!table) return { error: 'No table' };

            const rows = table.querySelectorAll('tr.el-table__row');
            if (rowIndex >= rows.length) return { error: 'Row not found: ' + rowIndex };

            const row = rows[rowIndex];
            // 获取行的所有文本
            const cells = row.querySelectorAll('td');
            const rowInfo = { rowIndex, cells: [] };
            cells.forEach((cell, ci) => {
                rowInfo.cells.push({
                    col: ci,
                    text: cell.innerText.trim().substring(0, 80),
                    imgs: Array.from(cell.querySelectorAll('img')).map(img => img.src)
                });
            });

            // 点击商品名称单元格（通常第2列）
            const titleCell = cells[1] || cells[0];
            titleCell.click();

            return rowInfo;
        }
    """, row_index)

    logger.info(f"[RowClick] Row {row_index}: {result}")
    return result


def fill_excel(product, template_path, output_path):
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["商品信息"]
    row = 3
    for r in range(1, 5):
        val = ws.cell(row=r, column=2).value
        if val and "商品标题" in str(val):
            row = r + 1; break
    for col_idx, value in enumerate(product.to_row(), start=1):
        ws.cell(row=row, column=col_idx, value=value)
    wb.save(output_path)


async def main(keyword: str):
    logger.info(f"=== ERP JS Extractor: {keyword} ===")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(viewport={"width": 1920, "height": 1080})

        # 加载 Cookie
        if COOKIES_FILE.exists():
            cookies = json.load(open(COOKIES_FILE, 'r', encoding='utf-8'))
            await context.add_cookies(cookies)
            logger.info(f"[Main] Loaded {len(cookies)} cookies")

        page = await context.new_page()

        # ========== Step 1: 导航到商品列表 ==========
        await page.goto(PRODUCT_LIST_URL, timeout=30000)
        await asyncio.sleep(6)  # 等待 Vue SPA 完整渲染

        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_js_list.png"), timeout=10000)

        # ========== Step 2: 探索页面结构 ==========
        info = await explore_page_structure(page)
        logger.info(f"[Explore] URL: {info['url']}")
        logger.info(f"[Explore] Title: {info['title']}")
        logger.info(f"[Explore] Visible inputs: {[i['placeholder'] for i in info['inputs']]}")
        logger.info(f"[Explore] elInputs: {info['elInputs']}")
        logger.info(f"[Explore] Buttons: {info['buttons']}")
        logger.info(f"[Explore] Row texts: {info['rowTexts']}")

        # ========== Step 3: 通过 JS 填充搜索 ==========
        await search_with_js(page, keyword)
        await asyncio.sleep(3)

        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_js_searched.png"), timeout=10000)

        # ========== Step 4: 获取搜索后的表格数据 ==========
        await asyncio.sleep(2)
        table_data = await get_table_data_via_js(page)
        logger.info(f"[Table] JS table data: {str(table_data)[:500]}")

        # ========== Step 5: 找匹配行 ==========
        matched_row_index = None
        if table_data.get('rows'):
            for row in table_data['rows']:
                row_text = ' '.join([c['text'] for c in row.get('cells', [])])
                if keyword.lower() in row_text.lower():
                    matched_row_index = row.get('index')
                    matched_cells = row.get('cells', [])
                    logger.info(f"[Match] Found at row {matched_row_index}: {row_text[:100]}")
                    break

        if matched_row_index is None:
            logger.warning(f"[Match] Keyword not found in table rows")
            logger.info(f"[Match] All rows: {[str(r)[:80] for r in table_data.get('rows', [])]}")

        # ========== Step 6: 构建商品数据 ==========
        product = ERPProduct()
        product.title = keyword

        if matched_row_index is not None and matched_cells:
            # 从匹配行提取
            for cell in matched_cells:
                col = cell['col']
                text = cell['text']
                imgs = cell.get('images', [])

                if col == 1:  # 商品名称
                    product.title = text
                    logger.info(f"[Data] Title: {text}")
                elif col == 2:  # 商品编码/SKU
                    product.sku_code = text
                elif col == 3:  # 分类
                    product.category = text
                elif col == 4:  # 价格相关
                    if '¥' in text or '￥' in text:
                        try:
                            price_num = float(re.findall(r'[\d.]+', text)[0])
                            if product.price == 99.0:
                                product.price = price_num
                            else:
                                product.original_price = price_num
                        except: pass
                elif col == 5:  # 原价
                    try:
                        product.original_price = float(re.findall(r'[\d.]+', text)[0])
                    except: pass
                elif col == 6:  # 库存
                    try:
                        product.stock = int(re.findall(r'\d+', text)[0])
                    except: pass
                elif col == 7:  # 是否上架
                    product.listed = 'Y' if '上架' in text or 'Y' in text else 'N'

                # 提取图片
                for img_url in imgs:
                    if 'logo' not in img_url.lower() and 'head' not in img_url.lower():
                        if not product.primary_image:
                            product.primary_image = img_url
                        else:
                            if product.other_images:
                                product.other_images += '\n' + img_url
                            else:
                                product.other_images = img_url

        # ========== Step 7: 点击进入详情页提取更多数据 ==========
        if matched_row_index is not None:
            try:
                # 点击该行
                click_result = await page.evaluate("""
                    (idx) => {
                        const table = document.querySelector('.el-table__body-wrapper');
                        if (!table) return 'no table';
                        const rows = table.querySelectorAll('tr.el-table__row');
                        if (idx >= rows.length) return 'no row';
                        // 点击第2个单元格（商品名称）
                        const cells = rows[idx].querySelectorAll('td');
                        if (cells[1]) cells[1].click();
                        else if (cells[0]) cells[0].click();
                        return 'clicked';
                    }
                """, matched_row_index)
                logger.info(f"[Detail] Click result: {click_result}")

                await asyncio.sleep(5)  # 等待详情页加载

                await page.screenshot(path=str(SCREENSHOT_DIR / "erp_js_detail.png"), timeout=10000)

                # 从详情页提取
                detail_data = await page.evaluate("""
                    () => {
                        const result = { url: window.location.href, forms: {} };

                        // TinyMCE iframe
                        const iframes = document.querySelectorAll('iframe');
                        for (const frame of iframes) {
                            try {
                                const frameDoc = frame.contentDocument || frame.contentWindow.document;
                                const body = frameDoc.body;
                                if (body && body.innerText) {
                                    result.tinymce = body.innerText.trim().substring(0, 200);
                                    break;
                                }
                            } catch(e) {}
                        }

                        // 所有图片
                        const imgs = [];
                        document.querySelectorAll('img').forEach(img => {
                            if (img.src && !img.src.includes('logo') && !img.src.includes('head')) {
                                imgs.push(img.src);
                            }
                        });
                        result.imgs = imgs.slice(0, 20);

                        // 价格输入框
                        const priceInputs = document.querySelectorAll('input');
                        priceInputs.forEach(inp => {
                            if (inp.placeholder && (inp.placeholder.includes('售价') || inp.placeholder.includes('原价') || inp.placeholder.includes('价格'))) {
                                result.forms[inp.placeholder] = inp.value;
                            }
                        });

                        // Vue 数据
                        try {
                            const app = document.querySelector('#app');
                            if (app && app.__vue__) {
                                result.vue = 'found';
                            }
                        } catch(e) { result.vue = e.message; }

                        return result;
                    }
                """)
                logger.info(f"[Detail] URL: {detail_data.get('url')}")
                logger.info(f"[Detail] TinyMCE: {detail_data.get('tinymce', 'not found')}")
                logger.info(f"[Detail] Images: {detail_data.get('imgs', [])[:3]}")
                logger.info(f"[Detail] Price inputs: {detail_data.get('forms', {})}")

                # 合并详情数据
                if detail_data.get('tinymce') and not product.description:
                    product.description = f"<p>{detail_data['tinymce']}</p>"

                if detail_data.get('imgs'):
                    existing_imgs = []
                    if product.primary_image:
                        existing_imgs = [product.primary_image]
                        if product.other_images:
                            existing_imgs.extend(product.other_images.split('\n'))
                    new_imgs = [img for img in detail_data['imgs']
                                if img not in existing_imgs and 'logo' not in img.lower()]
                    if not product.primary_image and new_imgs:
                        product.primary_image = new_imgs[0]
                        product.other_images = '\n'.join(new_imgs[1:14])

            except Exception as e:
                logger.warning(f"[Detail] Extraction error: {e}")

        # SEO Handle
        product.seo_handle = re.sub(r'[^a-z0-9]+', '-', product.title.lower())
        product.seo_handle = re.sub(r'^-+|-+$', '', product.seo_handle)[:100]

        # ========== Step 8: 写 Excel ==========
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = re.sub(r'[\\/:*?"<>|]', '', keyword[:20])
        output_path = OUTPUT_DIR / f"ERP提取_{safe_title}_{ts}.xlsx"
        fill_excel(product, TEMPLATE_FILE, output_path)

        logger.info(f"=== DONE ===")
        logger.info(f"Title: {product.title}")
        logger.info(f"Price: {product.price} / Original: {product.original_price}")
        logger.info(f"Primary Image: {product.primary_image[:80] if product.primary_image else 'NONE'}")
        logger.info(f"Other Images ({len(product.other_images.split(chr(10))) if product.other_images else 0}): ...")
        logger.info(f"Listed: {product.listed}")
        logger.info(f"Output: {output_path}")

        await asyncio.sleep(2)
        await browser.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("keyword", help="商品关键词")
    args = parser.parse_args()
    asyncio.run(main(args.keyword))
