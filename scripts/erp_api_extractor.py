#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP API 拦截抓取脚本 - 拦截商品列表 API，提取商品详情
API Reverse Extractor: Intercept ERP API to get product detail data

Usage:
    python scripts/erp_api_extractor.py "Descente Training Short Sleeve Polo Black"
"""

import asyncio
import json
import logging
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import async_playwright, BrowserContext, Page

# =============================================================================
# 路径配置
# =============================================================================
ROOT_DIR = Path(__file__).parent.parent
COOKIES_FILE = ROOT_DIR / "logs" / "cookies.json"
TEMPLATE_FILE = ROOT_DIR / "商品导入模板.xlsx"
OUTPUT_DIR = ROOT_DIR / "output"
SCREENSHOT_DIR = ROOT_DIR / "screenshots"
LOG_DIR = ROOT_DIR / "logs"

OUTPUT_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
SCREENSHOT_DIR.mkdir(exist_ok=True)

# =============================================================================
# 日志
# =============================================================================
logger = logging.getLogger("erp_api")
logger.setLevel(logging.INFO)
fh = logging.FileHandler(LOG_DIR / f"erp_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log", encoding='utf-8')
fh.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s'))
logger.addHandler(fh)
sh = logging.StreamHandler()
logger.addHandler(sh)

ERP_BASE = "https://www.mrshopplus.com"
PRODUCT_LIST_URL = f"{ERP_BASE}/#/product/list_DTB_proProduct"

# =============================================================================
# 数据模型
# =============================================================================
class ERPProduct:
    def __init__(self):
        self.product_id = ""
        self.title = ""
        self.subtitle = ""
        self.description = ""
        self.primary_image = ""
        self.other_images = ""
        self.key_info = ""
        self.attributes = ""
        self.listed = "N"
        self.shipping_template = "默认模板"
        self.category = ""
        self.tags = ""
        self.unit = "件/个"
        self.notes = ""
        self.no_inventory = "Y"
        self.weight = 0.5
        self.pack_length = 0.0
        self.pack_width = 0.0
        self.pack_height = 0.0
        self.seo_title = ""
        self.seo_description = ""
        self.seo_keywords = ""
        self.seo_handle = ""
        self.spec1_name = ""
        self.spec2_name = ""
        self.spec3_name = ""
        self.spec4_name = ""
        self.sku_values = ""
        self.sku_images = ""
        self.price = 99.0
        self.original_price = 199.0
        self.stock = 100
        self.sku_code = ""

    def to_row(self):
        return [
            self.product_id, self.title, self.subtitle, self.description,
            self.primary_image, self.other_images, self.key_info, self.attributes,
            self.listed, self.shipping_template, self.category, self.tags, self.unit,
            self.notes, self.no_inventory, self.weight, self.pack_length,
            self.pack_width, self.pack_height, self.seo_title, self.seo_description,
            self.seo_keywords, self.seo_handle, self.spec1_name, self.spec2_name,
            self.spec3_name, self.spec4_name, self.sku_values, self.sku_images,
            self.price, self.original_price, self.stock, self.sku_code
        ]


def make_url_handle(title: str) -> str:
    """生成 URL 安全句柄"""
    handle = title.lower()
    handle = re.sub(r'[^a-z0-9]+', '-', handle)
    handle = re.sub(r'^-+|-+$', '', handle)
    return handle[:100]


async def intercept_api_and_search(page: Page, keyword: str) -> dict:
    """拦截商品列表 API，搜索关键词"""

    api_data = {"rows": [], "total": 0}
    api_captured = asyncio.Event()

    async def handle_response(response):
        """捕获商品列表 API 响应"""
        if "/proProduct/page" in response.url or "/proProduct/list" in response.url:
            try:
                json_data = await response.json()
                # 打印数据结构用于调试
                logger.info(f"[API] Intercepted: {response.url[:100]}")
                logger.info(f"[API] Response keys: {list(json_data.keys())}")

                # 尝试找数据
                if "data" in json_data:
                    if isinstance(json_data["data"], dict):
                        records = json_data["data"].get("records", [])
                        api_data["rows"] = records
                        api_data["total"] = json_data["data"].get("total", 0)
                        logger.info(f"[API] Found {len(records)} records, total={api_data['total']}")
                    elif isinstance(json_data["data"], list):
                        api_data["rows"] = json_data["data"]
                        api_data["total"] = len(json_data["data"])

                if not api_captured.is_set():
                    api_captured.set()
            except Exception as e:
                logger.warning(f"[API] Parse error: {e}")

    # 注册监听
    page.on("response", handle_response)

    # 访问商品列表
    logger.info("[Search] Navigating to product list...")
    await page.goto(PRODUCT_LIST_URL, timeout=30000)
    await asyncio.sleep(3)

    # 触发搜索
    search_selector = ".el-input__inner[placeholder*='搜索']"
    try:
        await page.wait_for_selector(search_selector, timeout=8000)
        await page.fill(search_selector, keyword)
        await asyncio.sleep(1)

        # 按 Enter 触发搜索
        await page.keyboard.press("Enter")
        logger.info(f"[Search] Filled keyword: {keyword}")
    except Exception as e:
        logger.warning(f"[Search] Fill failed: {e}")
        # 尝试其他方式
        try:
            inputs = await page.query_selector_all("input")
            for inp in inputs:
                ph = await inp.get_attribute("placeholder") or ""
                if "搜索" in ph or "商品" in ph:
                    await inp.fill(keyword)
                    await page.keyboard.press("Enter")
                    logger.info(f"[Search] Filled via alt selector: {ph}")
                    break
        except Exception as e2:
            logger.warning(f"[Search] Alt fill failed: {e2}")

    # 等待 API 响应（最多8秒）
    try:
        await asyncio.wait_for(api_captured.wait(), timeout=8)
    except asyncio.TimeoutError:
        logger.warning("[Search] API timeout, checking captured data...")

    # 打印抓到的数据
    if api_data["rows"]:
        logger.info(f"[Search] Captured {len(api_data['rows'])} rows from API")
        for i, row in enumerate(api_data["rows"]):
            logger.info(f"[Search] Row {i}: {str(row)[:200]}")
    else:
        logger.warning("[Search] No API rows captured")

    # 截图
    await page.screenshot(path=str(SCREENSHOT_DIR / "erp_api_search.png"), timeout=10000)

    return api_data


async def get_product_detail(page: Page, product_id: str, pk_values: str) -> Optional[ERPProduct]:
    """通过编辑页 URL 访问商品详情"""

    # ERP 编辑页 URL 格式: /#/product/form_DTB_proProduct/0?action=4&pkValues=xxx
    detail_url = f"{ERP_BASE}/#/product/form_DTB_proProduct/0?action=4&pkValues={pk_values}"
    logger.info(f"[Detail] Navigating to: {detail_url}")

    await page.goto(detail_url, timeout=30000)
    await asyncio.sleep(5)  # 等待 Vue 组件加载

    await page.screenshot(path=str(SCREENSHOT_DIR / "erp_detail_page.png"), timeout=10000)

    product = ERPProduct()

    # 尝试从 API 拦截详情数据
    detail_data = {}

    def handle_response(resp):
        try:
            if "/proProduct/" in resp.url and ("detail" in resp.url or "get" in resp.url or resp.url.endswith(product_id)):
                logger.info(f"[Detail API] {resp.url}")
        except:
            pass

    page.on("response", handle_response)

    # 从 DOM 提取字段
    try:
        body_text = await page.evaluate("document.body.innerText")
        logger.info(f"[Detail] Page text preview: {body_text[:300]}")
    except:
        pass

    # 提取 TinyMCE iframe 内容
    try:
        for frame in page.frames:
            try:
                body = await frame.query_selector("#tinymce, .tox-edit-area__iframe")
                if body:
                    desc_text = await body.inner_text()
                    product.description = f"<p>{desc_text.strip()}</p>"
                    logger.info(f"[Detail] TinyMCE text: {desc_text[:50]}")
            except:
                continue
    except Exception as e:
        logger.warning(f"[Detail] TinyMCE extract: {e}")

    # 提取图片
    try:
        img_elements = await page.query_selector_all("img[src*='images.mrshopplus'], img[src*='oss']")
        img_urls = []
        for img in img_elements:
            src = await img.get_attribute("src")
            if src and "logo" not in src.lower() and "head" not in src.lower():
                img_urls.append(src)
        if img_urls:
            product.primary_image = img_urls[0]
            product.other_images = "\n".join(img_urls[1:14])
            logger.info(f"[Detail] Images: {len(img_urls)} total")
    except Exception as e:
        logger.warning(f"[Detail] Image extract: {e}")

    return product


def fill_excel_template(product: ERPProduct, template_path: Path, output_path: Path):
    """填入 Excel 模板"""
    import shutil
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb["商品信息"]

    # 找到数据行（表头在第2行，数据从第3行开始）
    row = 3
    for r in range(1, 5):
        val = ws.cell(row=r, column=2).value
        if val and "商品标题" in str(val):
            row = r + 1
            break

    for col_idx, value in enumerate(product.to_row(), start=1):
        ws.cell(row=row, column=col_idx, value=value)

    wb.save(output_path)
    logger.info(f"[Excel] Saved: {output_path}")


async def main(keyword: str):
    logger.info(f"{'='*60}")
    logger.info(f"[Main] ERP API Extractor - keyword: {keyword}")
    logger.info(f"{'='*60}")

    if not TEMPLATE_FILE.exists():
        logger.error(f"[Main] Template not found: {TEMPLATE_FILE}")
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )

        # 加载 Cookie
        if COOKIES_FILE.exists():
            with open(COOKIES_FILE, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)
            logger.info(f"[Main] Loaded {len(cookies)} cookies")

        page = await context.new_page()

        # 拦截 API 搜索
        api_data = await intercept_api_and_search(page, keyword)

        product = None

        if api_data["rows"]:
            # 找到匹配的商品
            matched = None
            for row in api_data["rows"]:
                title_in_row = str(row.get("productName", row.get("title", "")))
                if keyword.lower() in title_in_row.lower():
                    matched = row
                    break

            if matched:
                logger.info(f"[Main] MATCHED: {matched}")

                product = ERPProduct()
                product.title = matched.get("productName", keyword)
                product.product_id = matched.get("id", "") or matched.get("productId", "")
                product.price = float(matched.get("price", matched.get("salePrice", 99)))
                product.original_price = float(matched.get("originalPrice", matched.get("original_price", 199)))
                product.stock = int(matched.get("stock", matched.get("inventory", 100)))
                product.listed = "Y" if matched.get("isListed", matched.get("isOnline", "")) == "Y" else "N"
                product.category = matched.get("categoryName", matched.get("category", ""))
                product.tags = matched.get("tags", "")
                product.weight = float(matched.get("weight", 0.5))

                # SEO Handle
                product.seo_handle = make_url_handle(product.title)

                # 图片
                primary = matched.get("primaryImage", matched.get("mainImage", ""))
                others = matched.get("otherImages", matched.get("imageUrls", ""))
                if isinstance(others, list):
                    others = "\n".join(others)

                if primary:
                    product.primary_image = primary
                    if others:
                        product.other_images = others
                elif isinstance(others, str) and others:
                    imgs = others.split("\n")
                    if imgs:
                        product.primary_image = imgs[0]
                        product.other_images = "\n".join(imgs[1:14])

                # 描述
                desc = matched.get("description", matched.get("productDesc", ""))
                if desc:
                    # 清除 img 标签
                    desc_clean = re.sub(r'<img[^>]*>', '', str(desc))
                    product.description = desc_clean

                # 尝试从详情页获取更多信息
                pk = matched.get("pkValues", matched.get("id", ""))
                if pk:
                    try:
                        detail = await get_product_detail(page, product.product_id, pk)
                        if detail:
                            # 合并详情数据
                            if detail.description and not product.description:
                                product.description = detail.description
                            if detail.primary_image and not product.primary_image:
                                product.primary_image = detail.primary_image
                                product.other_images = detail.other_images
                    except Exception as e:
                        logger.warning(f"[Main] Detail fetch failed: {e}")

            else:
                logger.warning(f"[Main] Keyword '{keyword}' not found in API results")
                logger.info(f"[Main] Available titles: {[str(r.get('productName',''))[:50] for r in api_data['rows'][:5]]}")

        if not product:
            product = ERPProduct()
            product.title = keyword
            product.description = "<p>数据未找到，请手动填写</p>"
            logger.warning("[Main] Using fallback empty product")

        # 写入 Excel
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"ERP提取_{keyword[:20]}_{ts}.xlsx"
        fill_excel_template(product, TEMPLATE_FILE, output_path)

        logger.info(f"{'='*60}")
        logger.info(f"[Main] DONE!")
        logger.info(f"[Main] Title: {product.title}")
        logger.info(f"[Main] Price: {product.price} / Original: {product.original_price}")
        logger.info(f"[Main] Primary Image: {product.primary_image[:80] if product.primary_image else '(none)'}")
        logger.info(f"[Main] Listed: {product.listed}")
        logger.info(f"[Main] Output: {output_path}")
        logger.info(f"{'='*60}")

        await asyncio.sleep(3)
        await context.close()
        await browser.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("keyword", help="商品关键词")
    args = parser.parse_args()
    asyncio.run(main(args.keyword))
