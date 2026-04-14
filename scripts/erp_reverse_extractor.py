#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP 逆向抓取脚本 - 从 MrShopPlus ERP 提取商品数据填入 Excel 模板
Reverse Extractor: Crawl product from ERP and fill into Excel template

Usage:
    python scripts/erp_reverse_extractor.py "Descente Training Short Sleeve Polo Black"
    python scripts/erp_reverse_extractor.py "Descente Training Short Sleeve Polo Black" --save-screenshot
"""

import asyncio
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional, List

# 第三方
import requests
from openpyxl import load_workbook
from playwright.async_api import async_playwright, Page, BrowserContext

# =============================================================================
# 路径配置
# =============================================================================
ROOT_DIR = Path(__file__).parent.parent
COOKIES_FILE = ROOT_DIR / "logs" / "cookies.json"
TEMPLATE_FILE = ROOT_DIR / "商品导入模板.xlsx"
OUTPUT_FILE = ROOT_DIR / "output" / "ERP提取_商品信息.xlsx"
SCREENSHOT_DIR = ROOT_DIR / "screenshots"
LOG_DIR = ROOT_DIR / "logs"

LOG_DIR.mkdir(exist_ok=True)
SCREENSHOT_DIR.mkdir(exist_ok=True)
Path(ROOT_DIR / "output").mkdir(exist_ok=True)

# =============================================================================
# 日志配置
# =============================================================================
logger = logging.getLogger("erp_reverse")
logger.setLevel(logging.INFO)
fh = logging.FileHandler(LOG_DIR / f"erp_reverse_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log", encoding='utf-8')
fh.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s'))
logger.addHandler(fh)
sh = logging.StreamHandler()
logger.addHandler(sh)

# =============================================================================
# 常量
# =============================================================================
ERP_BASE = "https://www.mrshopplus.com"
PRODUCT_LIST_URL = f"{ERP_BASE}/#/product/list_DTB_proProduct"
LOGIN_URL = f"{ERP_BASE}/#/login"


# =============================================================================
# 数据模型
# =============================================================================
@dataclass
class ERPProduct:
    """ERP 商品数据（33列对应 Excel 模板）"""
    # A 商品ID
    product_id: str = ""
    # B 商品标题
    title: str = ""
    # C 副标题
    subtitle: str = ""
    # D 商品描述 (HTML)
    description: str = ""
    # E 商品首图 URL
    primary_image: str = ""
    # F 商品其他图片 (换行分隔)
    other_images: str = ""
    # G 关键信息 (HTML)
    key_info: str = ""
    # H 属性 (属性名|属性值)
    attributes: str = ""
    # I 商品上架 (Y/N)
    listed: str = "N"
    # J 物流模板
    shipping_template: str = "默认模板"
    # K 类别名称
    category: str = ""
    # L 标签
    tags: str = ""
    # M 计量单位
    unit: str = "件/个"
    # N 商品备注
    notes: str = ""
    # O 不记库存 (Y/N)
    no_inventory: str = "Y"
    # P 商品重量 (kg)
    weight: float = 0.5
    # Q 包装长度 (cm)
    pack_length: float = 0
    # R 包装宽度 (cm)
    pack_width: float = 0
    # S 包装高度 (cm)
    pack_height: float = 0
    # T SEO标题
    seo_title: str = ""
    # U SEO描述
    seo_description: str = ""
    # V SEO关键词
    seo_keywords: str = ""
    # W SEO URL Handle
    seo_handle: str = ""
    # X 规格1
    spec1_name: str = ""
    # Y 规格2
    spec2_name: str = ""
    # Z 规格3
    spec3_name: str = ""
    # AA 规格4
    spec4_name: str = ""
    # AB SKU值 (Color:Yellow\nSize:41)
    sku_values: str = ""
    # AC SKU图片 URL
    sku_images: str = ""
    # AD 售价
    price: float = 99.0
    # AE 原价
    original_price: float = 199.0
    # AF 库存
    stock: int = 100
    # AG SKU编码
    sku_code: str = ""

    def to_row(self) -> list:
        """转为 Excel 行（33列，与模板顺序一致）"""
        return [
            self.product_id,      # A
            self.title,           # B
            self.subtitle,        # C
            self.description,     # D
            self.primary_image,   # E
            self.other_images,    # F
            self.key_info,        # G
            self.attributes,      # H
            self.listed,          # I
            self.shipping_template, # J
            self.category,        # K
            self.tags,            # L
            self.unit,            # M
            self.notes,           # N
            self.no_inventory,    # O
            self.weight,          # P
            self.pack_length,     # Q
            self.pack_width,      # R
            self.pack_height,     # S
            self.seo_title,       # T
            self.seo_description, # U
            self.seo_keywords,    # V
            self.seo_handle,      # W
            self.spec1_name,      # X
            self.spec2_name,      # Y
            self.spec3_name,      # Z
            self.spec4_name,      # AA
            self.sku_values,      # AB
            self.sku_images,      # AC
            self.price,           # AD
            self.original_price,  # AE
            self.stock,           # AF
            self.sku_code,        # AG
        ]


# =============================================================================
# ERP 登录 & 数据抓取
# =============================================================================

async def erp_login(context: BrowserContext, cookies_file: Path) -> Optional[Page]:
    """登录 ERP（优先 Cookie 注入，兜底 UI 登录）"""
    page = await context.new_page()

    # Step 1: 尝试加载 Cookie
    if cookies_file.exists():
        try:
            with open(cookies_file, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            await context.add_cookies(cookies)
            logger.info(f"[Login] Loaded {len(cookies)} cookies from {cookies_file}")

            # 验证 Cookie 是否有效
            await page.goto(PRODUCT_LIST_URL, timeout=20000)
            await asyncio.sleep(2)

            # 检查是否被重定向到登录页
            if "login" in page.url or page.url == ERP_BASE + "/":
                logger.info("[Login] Cookies expired, need re-login")
                # 清除过期 Cookie 重新登录
                await context.clear_cookies()
            else:
                logger.info(f"[Login] Cookie OK, current URL: {page.url}")
                return page
        except Exception as e:
            logger.warning(f"[Login] Cookie load failed: {e}")

    # Step 2: UI 登录
    await page.goto(LOGIN_URL, timeout=20000)
    await asyncio.sleep(3)

    # 填充登录表单
    await page.fill("#username", "zhiqiang")
    await page.fill("input[placeholder='请输入密码']", "123qazwsx")
    await page.click("#login-btn")
    await asyncio.sleep(5)

    if "login" in page.url:
        logger.error("[Login] UI login failed")
        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_login_failed.png"))
        await page.close()
        return None

    # 保存新 Cookie
    new_cookies = await context.cookies()
    with open(cookies_file, 'w', encoding='utf-8') as f:
        json.dump(new_cookies, f, f, indent=2)
    logger.info(f"[Login] UI login success, saved {len(new_cookies)} cookies")

    return page


async def search_and_extract(page: Page, keyword: str) -> Optional[ERPProduct]:
    """在 ERP 商品列表搜索关键词，提取商品数据"""

    logger.info(f"[Search] Navigating to product list for keyword: {keyword}")

    # 导航到商品列表
    await page.goto(PRODUCT_LIST_URL, timeout=30000)
    await asyncio.sleep(5)  # Vue SPA 渲染等待

    # ========== 调试：打印当前页面状态 ==========
    logger.info(f"[Debug] Current URL: {page.url}")
    title = await page.title()
    logger.info(f"[Debug] Page title: {title}")

    # 截图用于调试
    await page.screenshot(path=str(SCREENSHOT_DIR / "erp_product_list.png"), timeout=10000)
    logger.info(f"[Debug] Screenshot saved: erp_product_list.png")

    # ========== 尝试在列表中搜索 ==========
    search_selector = ".el-input__inner[placeholder*='搜索']"
    try:
        await page.wait_for_selector(search_selector, timeout=10000)
        await page.fill(search_selector, keyword)
        await asyncio.sleep(2)
        logger.info("[Search] Search input filled")
    except Exception as e:
        logger.warning(f"[Search] Search input not found: {e}")
        # 尝试其他常见搜索选择器
        for selector in [
            "input[placeholder='搜索']",
            ".el-search-input input",
            "input[placeholder*='商品']",
        ]:
            try:
                el = page.locator(selector).first
                if await el.is_visible():
                    await el.fill(keyword)
                    logger.info(f"[Search] Found search with selector: {selector}")
                    await asyncio.sleep(2)
                    break
            except:
                continue

    # 再次截图
    await page.screenshot(path=str(SCREENSHOT_DIR / "erp_search_result.png"), timeout=10000)
    logger.info(f"[Debug] Search result screenshot saved")

    # ========== 尝试点击搜索按钮或回车 ==========
    try:
        # 找搜索相关的按钮
        btns = await page.query_selector_all("button")
        for btn in btns:
            btn_text = await btn.inner_text()
            if any(k in btn_text for k in ["搜索", "查询", "Search", "搜索商品"]):
                await btn.click()
                logger.info(f"[Search] Clicked search button: {btn_text}")
                break
        else:
            # 按回车触发搜索
            await page.keyboard.press("Enter")
            logger.info("[Search] Pressed Enter")
    except Exception as e:
        logger.warning(f"[Search] Search trigger failed: {e}")

    await asyncio.sleep(3)

    # ========== 提取商品行 ==========
    product_data = await extract_product_from_list(page, keyword)
    if product_data:
        return product_data

    # ========== 如果列表没找到，尝试直接访问商品详情 ==========
    logger.info("[Search] Not found in list, trying detail page approach")
    return await extract_from_detail_page(page, keyword)


async def extract_product_from_list(page: Page, keyword: str) -> Optional[ERPProduct]:
    """从商品列表提取数据"""
    try:
        # 等待列表加载
        await asyncio.sleep(3)

        # Vue Element Plus 表格行选择器
        row_selector = ".el-table__body-wrapper tr.el-table__row"
        rows = await page.query_selector_all(row_selector)

        logger.info(f"[Extract] Found {len(rows)} table rows")

        if not rows:
            # 尝试其他选择器
            for selector in ["tr", ".el-table__row"]:
                rows = await page.query_selector_all(selector)
                if rows:
                    logger.info(f"[Extract] Found {len(rows)} rows with: {selector}")
                    break

        for i, row in enumerate(rows):
            # 提取行内文字
            row_text = await row.inner_text()
            logger.info(f"[Extract] Row {i} text preview: {row_text[:100]}")

            if keyword.lower() in row_text.lower():
                logger.info(f"[Extract] Matched row {i} for keyword: {keyword}")

                # 点击该行进入详情
                try:
                    # 查找商品标题单元格并点击
                    title_cell = row.query_selector("td:nth-child(2)")
                    if title_cell:
                        await title_cell.click()
                        logger.info(f"[Extract] Clicked title cell")
                        await asyncio.sleep(4)

                        # 截图详情页
                        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_product_detail.png"), timeout=10000)
                        logger.info(f"[Extract] Detail page screenshot saved")

                        # 从详情页提取数据
                        product = await extract_from_detail_form(page, keyword)
                        if product:
                            return product
                except Exception as e:
                    logger.warning(f"[Extract] Row click failed: {e}")

        return None

    except Exception as e:
        logger.error(f"[Extract] List extraction failed: {e}")
        return None


async def extract_from_detail_form(page: Page, keyword: str) -> Optional[ERPProduct]:
    """从商品编辑详情表单提取所有字段"""
    try:
        await asyncio.sleep(3)  # 等待表单渲染

        product = ERPProduct()
        product.title = keyword  # 已知关键词作为标题

        # 截图
        await page.screenshot(path=str(SCREENSHOT_DIR / "erp_detail_form.png"), timeout=10000)

        # ========== 提取表单字段 ==========

        # 商品标题
        try:
            title_input = page.locator("input[placeholder*='商品名称'], input[placeholder*='标题']").first
            if await title_input.is_visible():
                product.title = await title_input.input_value()
                logger.info(f"[Extract] Title: {product.title}")
        except Exception as e:
            logger.warning(f"[Extract] Title extraction: {e}")

        # 商品描述 (TinyMCE iframe)
        try:
            # TinyMCE 编辑器在 iframe 中
            frames = page.frames
            logger.info(f"[Extract] Found {len(frames)} frames")
            for frame in frames:
                try:
                    body = await frame.query_selector("#tinymce")
                    if body:
                        desc = await body.inner_text()
                        product.description = f"<p>{desc.strip()}</p>"
                        logger.info(f"[Extract] Description from TinyMCE: {desc[:50]}")
                        break
                except:
                    continue
        except Exception as e:
            logger.warning(f"[Extract] Description extraction: {e}")

        # 商品图片 - 提取页面中所有图片
        try:
            imgs = await page.query_selector_all("img[src*='images.mrshopplus'], img[src*='image']")
            img_urls = []
            for img in imgs:
                src = await img.get_attribute("src")
                if src and "logo" not in src.lower():
                    img_urls.append(src)

            if img_urls:
                product.primary_image = img_urls[0]
                product.other_images = "\n".join(img_urls[1:14])
                logger.info(f"[Extract] Images: primary={product.primary_image}, others={len(img_urls)-1}")
        except Exception as e:
            logger.warning(f"[Extract] Images extraction: {e}")

        # 商品价格
        try:
            price_selectors = [
                "input[placeholder*='售价']",
                "input[placeholder*='价格']",
                ".el-input-number input"
            ]
            for sel in price_selectors:
                inputs = await page.query_selector_all(sel)
                for inp in inputs:
                    if await inp.is_visible():
                        val = await inp.input_value()
                        try:
                            price_val = float(val)
                            if price_val > 0:
                                if product.price == 99.0:  # 默认值未改
                                    product.price = price_val
                                    logger.info(f"[Extract] Price: {product.price}")
                                elif product.original_price == 199.0:
                                    product.original_price = price_val
                                    logger.info(f"[Extract] Original price: {product.original_price}")
                        except:
                            pass
                        break
        except Exception as e:
            logger.warning(f"[Extract] Price extraction: {e}")

        # SKU相关信息
        try:
            sku_selectors = [
                "input[placeholder*='SKU']",
                ".el-table input[placeholder*='规格']"
            ]
            for sel in sku_selectors:
                inputs = await page.query_selector_all(sel)
                for inp in inputs:
                    if await inp.is_visible():
                        val = await inp.input_value()
                        if val:
                            product.sku_code = val
                            logger.info(f"[Extract] SKU: {product.sku_code}")
                            break
        except Exception as e:
            logger.warning(f"[Extract] SKU extraction: {e}")

        # 分类
        try:
            cat_selectors = [
                ".el-select[placeholder*='类别']",
                "input[placeholder*='类别']",
                ".el-cascader input"
            ]
            for sel in cat_selectors:
                el = page.locator(sel).first
                if await el.is_visible():
                    val = await el.input_value()
                    if val:
                        product.category = val
                        logger.info(f"[Extract] Category: {product.category}")
                        break
        except Exception as e:
            logger.warning(f"[Extract] Category extraction: {e}")

        # 标签
        try:
            tag_el = page.locator(".el-tag, input[placeholder*='标签']").first
            if await tag_el.is_visible():
                product.tags = await tag_el.inner_text()
                logger.info(f"[Extract] Tags: {product.tags}")
        except:
            pass

        # 提取页面中所有文字内容用于调试
        try:
            body_text = await page.evaluate("document.body.innerText")
            logger.info(f"[Debug] Page text preview: {body_text[:300]}")
        except Exception as e:
            logger.warning(f"[Debug] Page text extraction: {e}")

        logger.info(f"[Extract] Final product data: title={product.title}, price={product.price}")
        return product

    except Exception as e:
        logger.error(f"[Extract] Detail form extraction failed: {e}")
        return None


async def extract_from_detail_page(page: Page, keyword: str) -> Optional[ERPProduct]:
    """兜底方案：从页面 DOM 提取商品数据"""
    logger.info("[Extract] Trying DOM-based extraction")

    try:
        # 获取页面所有文本
        all_text = await page.evaluate("document.body.innerText")
        logger.info(f"[Extract] Page text length: {len(all_text)}")

        product = ERPProduct()
        product.title = keyword

        # 尝试正则匹配价格
        price_matches = re.findall(r'(\d+\.?\d*)\s*(?:USD|原价|售价|$)', all_text)
        if price_matches:
            try:
                product.price = float(price_matches[0])
                logger.info(f"[Extract] Price from text: {product.price}")
            except:
                pass

        # 尝试匹配图片URL
        img_matches = re.findall(r'https?://[^\s"\'<>]+\.(?:jpg|jpeg|png|webp)', all_text)
        if img_matches:
            product.primary_image = img_matches[0]
            product.other_images = "\n".join(img_matches[1:14])
            logger.info(f"[Extract] Found {len(img_matches)} image URLs")

        return product

    except Exception as e:
        logger.error(f"[Extract] DOM extraction failed: {e}")
        return None


# =============================================================================
# Excel 写入
# =============================================================================

def fill_excel_template(product: ERPProduct, template_path: Path, output_path: Path):
    """将商品数据填入 Excel 模板"""

    logger.info(f"[Excel] Loading template: {template_path}")

    # 复制模板
    import shutil
    shutil.copy(template_path, output_path)

    # 打开 Excel
    wb = load_workbook(output_path)
    ws = wb["商品信息"]

    # 商品信息 sheet 在第3行开始是数据（前2行是表头说明）
    row = ws.max_row + 1

    # 找到数据开始的行（跳过说明行）
    for r in range(1, 10):
        cell_val = ws.cell(row=r, column=2).value  # B列是商品标题
        if cell_val and "商品标题" in str(cell_val):
            row = r + 1
            logger.info(f"[Excel] Data header found at row {r}, data starts row {row}")
            break

    # 写入商品数据（33列）
    data_row = product.to_row()
    for col_idx, value in enumerate(data_row, start=1):
        ws.cell(row=row, column=col_idx, value=value)

    wb.save(output_path)
    logger.info(f"[Excel] Saved to: {output_path}")
    return output_path


# =============================================================================
# 主流程
# =============================================================================

async def main(keyword: str, save_screenshot: bool = False):
    """逆向抓取 ERP 商品数据"""

    logger.info(f"=" * 60)
    logger.info(f"[Main] Starting ERP Reverse Extractor")
    logger.info(f"[Main] Keyword: {keyword}")
    logger.info(f"[Main] Template: {TEMPLATE_FILE}")
    logger.info(f"=" * 60)

    # 验证模板存在
    if not TEMPLATE_FILE.exists():
        logger.error(f"[Main] Template not found: {TEMPLATE_FILE}")
        return

    # 启动浏览器
    async with async_playwright() as p:
        # 无头模式（headless=False 可看到浏览器）
        browser = await p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"]
        )

        # 创建独立上下文（不共享其他会话）
        context = await browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )

        try:
            # Step 1: 登录 ERP
            page = await erp_login(context, COOKIES_FILE)
            if not page:
                logger.error("[Main] ERP login failed, aborting")
                return

            # Step 2: 搜索并提取商品
            product = await search_and_extract(page, keyword)
            if not product:
                logger.warning("[Main] Product not found or extraction failed")
                # 保存兜底数据
                product = ERPProduct()
                product.title = keyword
                product.description = "<p>数据提取失败，请手动填写</p>"
                logger.info("[Main] Using fallback product data")

            # Step 3: 写入 Excel
            output_path = fill_excel_template(product, TEMPLATE_FILE, OUTPUT_FILE)

            logger.info("=" * 60)
            logger.info(f"[Main] DONE! Output: {output_path}")
            logger.info(f"[Main] Product title: {product.title}")
            logger.info(f"[Main] Primary image: {product.primary_image}")
            logger.info(f"[Main] Price: {product.price} / Original: {product.original_price}")
            logger.info("=" * 60)

        finally:
            await context.close()
            await browser.close()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="ERP Reverse Extractor")
    parser.add_argument("keyword", help="商品关键词/标题")
    parser.add_argument("--screenshot", action="store_true", help="保存截图")
    args = parser.parse_args()

    asyncio.run(main(args.keyword, args.screenshot))
