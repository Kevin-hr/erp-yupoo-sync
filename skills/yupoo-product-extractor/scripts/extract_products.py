#!/usr/bin/env python3
"""
Yupoo Product Extractor - Extract products from Yupoo albums and export to Excel

Usage:
    python skills/yupoo-product-extractor/scripts/extract_products.py --date yesterday
    python skills/yupoo-product-extractor/scripts/extract_products.py --date 2026-04-21
    python skills/yupoo-product-extractor/scripts/extract_products.py --days 7
"""

import argparse
import json
import re
import subprocess
import sys
import time
import uuid
import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# Configuration
YUPOO_BASE_URL = "https://lol2024.x.yupoo.com/albums"
OUTPUT_DIR = Path("inputs")
PARSER_SCRIPT = r"C:\Users\Administrator\AppData\Roaming\qianfan-desktop-app\qianfan_desk_xdg\global\data\skills\baidu-content-parser\scripts\parse.py"

# English product name translations
ENGLISH_NAMES = {
    "Louis Vuitton": "Louis Vuitton",
    "Burberry": "Burberry",
    "Loewe": "Loewe",
    "MIU MIU": "MIU MIU",
    "GUCCI": "GUCCI",
    "Gucci": "Gucci",
    "Moncler": "Moncler",
    "Balenciaga": "Balenciaga",
    "Givenchy": "Givenchy",
    "Celine": "Celine",
    "Dior": "Dior",
    "OFF-WHITE": "OFF-WHITE",
    "Thom Browne": "Thom Browne",
    "PRADA": "PRADA",
    "Alexander Wang": "Alexander Wang",
    "Ami Paris": "Ami Paris",
}

# Color patterns
COLOR_PATTERNS = {
    "White": [r"白色", r"白款", r"white", r"White"],
    "Black": [r"黑色", r"黑款", r"black", r"Black"],
    "Gray": [r"灰色", r"灰款", r"gray", r"grey", r"Gray"],
    "Blue": [r"蓝色", r"蓝款", r"blue", r"Blue", r"深蓝", r"navy"],
    "Red": [r"红色", r"红款", r"red", r"Red"],
    "Green": [r"绿色", r"绿款", r"green", r"Green"],
    "Yellow": [r"黄色", r"黄款", r"yellow", r"Yellow"],
    "Pink": [r"粉色", r"粉款", r"pink", r"Pink"],
    "Orange": [r"橙色", r"橙款", r"orange", r"Orange"],
    "Brown": [r"棕色", r"棕款", r"brown", r"Brown", r"咖啡"],
    "Beige": [r"米色", r"米白", r"beige", r"Beige"],
    "Khaki": [r"卡其", r"khaki"],
}


def run_cli_command(cmd: str, timeout: int = 60) -> str:
    """Run playwright-cli command"""
    result = subprocess.run(
        f"playwright-cli {cmd}",
        shell=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
    )
    return result.stdout


def get_date_arg(date_str: str) -> str:
    """Convert date argument to YYYY-MM-DD format"""
    if date_str == "yesterday":
        return (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    return date_str


def extract_product_list() -> list:
    """Extract product list from Yupoo albums page"""
    print("Extracting product list from Yupoo...")

    # Navigate to albums page
    run_cli_command(f'goto "{YUPOO_BASE_URL}"')
    time.sleep(3)

    # Extract products from "昨天" section
    js_code = """(() => {
        const yesterdaySection = Array.from(document.querySelectorAll('main h2')).find(h => h.textContent.includes('昨天'))?.parentElement;
        if (!yesterdaySection) return [];
        const links = yesterdaySection.querySelectorAll('a[href*="/albums/"]');
        return Array.from(links).map(a => {
            const parent = a.parentElement;
            const nameEl = parent ? parent.querySelector('div:last-child') : null;
            return {
                name: a.getAttribute('title') || a.getAttribute('aria-label') || (nameEl ? nameEl.textContent.trim() : a.textContent.trim()),
                url: a.href,
                album_id: a.href.match(/albums\\/(\\d+)/)?.[1]
            };
        });
    })()"""

    result = run_cli_command(f'eval "{js_code}"')

    # Parse JSON
    try:
        start = result.find("[")
        end = result.rfind("]") + 1
        if start >= 0 and end > start:
            return json.loads(result[start:end])
    except Exception as e:
        print(f"Error parsing product list: {e}")

    return []


def extract_images_from_page() -> list:
    """Extract image URLs from current product page"""
    js_code = """(() => {
        const images = document.querySelectorAll('main img');
        return Array.from(images).map(img => {
            const originSrc = img.getAttribute('data-origin-src');
            return originSrc ? originSrc.replace('https://photo.yupoo.com', 'http://pic.yupoo.com') : null;
        }).filter(src => src);
    })()"""

    result = run_cli_command(f'eval "{js_code}"')

    try:
        start = result.find("[")
        end = result.rfind("]") + 1
        if start >= 0 and end > start:
            return json.loads(result[start:end])
    except:
        pass

    return []


def generate_english_name(chinese_name: str) -> str:
    """Generate English product name from Chinese name"""
    # Remove price prefix (h130, h140, etc.)
    name = re.sub(r"^[hH]?\d+", "", chinese_name).strip()

    # Find brand
    brand = None
    for brand_cn, brand_en in ENGLISH_NAMES.items():
        if brand_cn in name or brand_en in name:
            brand = brand_en
            break

    # Build English name
    if brand:
        # Extract product type
        if "短袖" in name or "T恤" in name:
            product_type = "T-Shirt"
        elif "短裤" in name:
            product_type = "Shorts"
        elif "外套" in name or "夹克" in name:
            product_type = "Jacket"
        elif "Polo" in name or "polo" in name:
            product_type = "Polo Shirt"
        else:
            product_type = "T-Shirt"

        # Build name
        english_name = f"{brand} {product_type}"
    else:
        # Default: use original name
        english_name = name

    return english_name


def extract_color_from_name(name: str) -> Optional[str]:
    """Extract color from product name"""
    for color, patterns in COLOR_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, name, re.IGNORECASE):
                return color
    return None


def create_excel(data: list, output_path: Path):
    """Create Excel file from product data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return False

    # Create temp directory
    tmp_dir = f".dumate/xlsx-{uuid.uuid4()}"
    os.makedirs(tmp_dir, exist_ok=True)

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Products"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "No.",
        "Product Name",
        "English Product Name",
        "Album ID",
        "Image Count",
        "First Image",
        "Second Image",
        "Image Links (Max 12)",
    ]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Column widths
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 50
    sheet.column_dimensions["C"].width = 65
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 10
    sheet.column_dimensions["F"].width = 50
    sheet.column_dimensions["G"].width = 50
    sheet.column_dimensions["H"].width = 80

    # Ensure unique English names
    english_names = set()

    # Add data
    row_num = 2
    for i, product in enumerate(data, 1):
        images = product.get("images", [])

        # Generate unique English name
        english_name = product.get(
            "english_name", generate_english_name(product["name"])
        )
        color = product.get("color") or extract_color_from_name(product["name"])

        if color:
            english_name = f"{english_name} - {color}"

        # Ensure uniqueness
        if english_name in english_names:
            counter = 2
            while f"{english_name} ({counter})" in english_names:
                counter += 1
            english_name = f"{english_name} ({counter})"
        english_names.add(english_name)

        # Add row
        sheet.cell(row=row_num, column=1, value=i).border = thin_border
        sheet.cell(row=row_num, column=2, value=product["name"]).border = thin_border
        sheet.cell(row=row_num, column=3, value=english_name).border = thin_border
        sheet.cell(
            row=row_num, column=4, value=product["album_id"]
        ).border = thin_border
        sheet.cell(row=row_num, column=5, value=len(images)).border = thin_border

        first_img = images[0] if images else ""
        sheet.cell(row=row_num, column=6, value=first_img).border = thin_border

        second_img = images[1] if len(images) > 1 else ""
        sheet.cell(row=row_num, column=7, value=second_img).border = thin_border

        remaining = images[2:14]
        cell = sheet.cell(row=row_num, column=8, value="\n".join(remaining))
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_num += 1

    # Save
    tmp_file = os.path.join(tmp_dir, "products.xlsx")
    wb.save(tmp_file)
    shutil.copy2(tmp_file, str(output_path))
    shutil.rmtree(tmp_dir, ignore_errors=True)

    return True


def main():
    parser = argparse.ArgumentParser(description="Extract products from Yupoo albums")
    parser.add_argument(
        "--date",
        type=str,
        default="yesterday",
        help='Date to extract (YYYY-MM-DD or "yesterday")',
    )
    parser.add_argument("--days", type=int, help="Extract products from last N days")
    parser.add_argument("--output", type=str, help="Output file path")
    args = parser.parse_args()

    # Determine date
    if args.days:
        date_str = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")
    else:
        date_str = get_date_arg(args.date)

    print(f"Extracting products for: {date_str}")

    # Output paths
    json_path = OUTPUT_DIR / f"yupoo_products_{date_str}.json"
    excel_path = OUTPUT_DIR / f"yupoo_products_{date_str}.xlsx"
    if args.output:
        excel_path = Path(args.output)

    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Extract product list
    products = extract_product_list()
    print(f"Found {len(products)} products")

    if not products:
        print("No products found. Make sure you're logged in to Yupoo.")
        return 1

    # Extract images for each product
    results = []
    for i, product in enumerate(products):
        print(f"[{i + 1}/{len(products)}] Extracting: {product['name'][:40]}...")

        # Navigate to product page
        run_cli_command(f'goto "{product["url"]}"')
        time.sleep(2)

        # Extract images
        images = extract_images_from_page()

        # Generate English name
        english_name = generate_english_name(product["name"])

        result = {
            **product,
            "images": images,
            "image_count": len(images),
            "english_name": english_name,
        }
        results.append(result)

        print(f"  Found {len(images)} images")

        # Save progress every 10 products
        if (i + 1) % 10 == 0:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)

    # Save final JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # Create Excel
    if create_excel(results, excel_path):
        print(f"\nCompleted!")
        print(f"JSON: {json_path}")
        print(f"Excel: {excel_path}")
        print(f"Total products: {len(results)}")
        print(f"Total images: {sum(r['image_count'] for r in results)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
